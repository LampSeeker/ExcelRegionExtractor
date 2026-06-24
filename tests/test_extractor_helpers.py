from __future__ import annotations

import json

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from excel_info_region.borders import boxes_are_contact_merge_neighbors
from excel_info_region.chart_export import extract_sheet_charts_to_dir
from excel_info_region.chart_regions import chart_boxes, chart_metadata
from excel_info_region.components import (
    connected_components_from_cells,
    remove_exact_or_contained_duplicates,
)
from excel_info_region.extractor import extract_info_regions_from_sheet, extract_workbook_info_regions
from excel_info_region.runner import _merge_structure_boxes, _region_tree, run_and_write
from excel_info_region.schema import Box


def test_connected_components_respects_connectivity():
    occupied = {(1, 1), (2, 2)}

    assert [b.range_ref for b in connected_components_from_cells(occupied, connectivity=4)] == [
        "A1:A1",
        "B2:B2",
    ]
    assert [b.range_ref for b in connected_components_from_cells(occupied, connectivity=8)] == [
        "A1:B2",
    ]


def test_border_contact_neighbors_need_gap_and_overlap():
    cfg = {
        "border_contact_merge_max_gap": 1,
        "border_contact_merge_min_axis_overlap": 0.8,
    }

    assert boxes_are_contact_merge_neighbors(Box(1, 1, 2, 4), Box(4, 1, 5, 4), cfg)
    assert not boxes_are_contact_merge_neighbors(Box(1, 1, 2, 4), Box(5, 1, 6, 4), cfg)
    assert not boxes_are_contact_merge_neighbors(Box(1, 1, 2, 4), Box(4, 4, 5, 7), cfg)
    assert not boxes_are_contact_merge_neighbors(Box(1, 1, 6, 6), Box(2, 2, 4, 4), cfg)


def test_remove_contained_duplicates_is_config_gated():
    boxes = [Box(1, 1, 4, 4), Box(2, 2, 3, 3), Box(1, 1, 4, 4)]

    assert remove_exact_or_contained_duplicates(boxes, {"remove_contained_duplicates": False}) == [
        Box(1, 1, 4, 4),
        Box(2, 2, 3, 3),
    ]
    assert remove_exact_or_contained_duplicates(boxes, {"remove_contained_duplicates": True}) == [
        Box(1, 1, 4, 4),
    ]


def test_respect_hidden_rows_cols_excludes_hidden_cells():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "visible"
    ws["B1"] = "hidden"
    ws.column_dimensions["B"].hidden = True

    result = extract_info_regions_from_sheet(
        ws,
        config={
            "respect_hidden_rows_cols": True,
            "include_images": False,
            "use_borders": False,
            "use_border_contact_merge": False,
        },
    )

    assert result["info_regions"] == ["A1:A1"]


def test_use_print_area_bounds_limits_regions():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "inside"
    ws["Z1"] = "outside"
    ws.print_area = "A1:B2"

    result = extract_info_regions_from_sheet(
        ws,
        config={
            "use_print_area_bounds": True,
            "include_images": False,
            "use_borders": False,
            "use_border_contact_merge": False,
        },
    )

    assert result["info_regions"] == ["A1:A1"]


def test_respect_hidden_sheets_skips_hidden_workbook_sheets(tmp_path):
    wb = Workbook()
    wb.active.title = "visible"
    hidden = wb.create_sheet("hidden")
    hidden.sheet_state = "hidden"
    path = tmp_path / "hidden_sheet.xlsx"
    wb.save(path)

    result = extract_workbook_info_regions(path, config={"respect_hidden_sheets": True})

    assert list(result["sheets"]) == ["visible"]


def test_region_tree_does_not_mark_contained_regions_as_children_without_border_evidence():
    assert _region_tree(["A1:P21", "L5:P14", "A22:J32"]) == [
        {"range_ref": "A1:P21", "children": []},
        {"range_ref": "A22:J32", "children": []},
    ]


def test_region_tree_adds_inner_bordered_tables_from_worksheet():
    wb = Workbook()
    ws = wb.active
    thin = "thin"
    from openpyxl.styles import Border, Side

    side = Side(style=thin)
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in range(3, 6):
        for col in range(3, 5):
            ws.cell(row, col).border = border

    assert _region_tree(["A1:F7"], ws) == [
        {
            "range_ref": "A1:F7",
            "children": [{"range_ref": "C3:D5"}],
        }
    ]


def test_region_tree_child_candidates_include_merged_value_cells():
    from openpyxl.styles import Border, Side

    wb = Workbook()
    ws = wb.active
    side = Side(style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in range(3, 6):
        for col in range(3, 6):
            ws.cell(row, col).border = border
    ws.merge_cells("C6:E6")
    ws["C6"] = "SUM"
    for col in range(3, 6):
        ws.cell(6, col).border = border

    assert _region_tree(["A1:G8"], ws) == [
        {"range_ref": "A1:G8", "children": [{"range_ref": "C3:E6"}]},
    ]


def test_region_tree_ignores_single_cell_contained_regions():
    wb = Workbook()
    ws = wb.active

    assert _region_tree(["A1:D35", "D8:D8"], ws) == [
        {"range_ref": "A1:D35", "children": []},
    ]


def test_region_tree_keeps_separate_roots_separate():
    from openpyxl.styles import Border, Side

    wb = Workbook()
    ws = wb.active
    side = Side(style="thin")
    full = Border(left=side, right=side, top=side, bottom=side)
    for row in range(1, 9):
        for col in range(1, 5):
            ws.cell(row, col).border = full

    assert _region_tree(["A1:D2", "A6:D8"], ws) == [
        {"range_ref": "A1:D2", "children": []},
        {"range_ref": "A6:D8", "children": []},
    ]

    wb = Workbook()
    ws = wb.active
    for row in [1, 2, 6, 7, 8]:
        for col in range(1, 5):
            ws.cell(row, col).border = full

    assert _region_tree(["A1:D2", "A6:D8"], ws) == [
        {"range_ref": "A1:D2", "children": []},
        {"range_ref": "A6:D8", "children": []},
    ]


def test_chart_metadata_includes_anchor_and_sources():
    wb = Workbook()
    ws = wb.active
    ws.append(["Label", "Score"])
    ws.append(["A", 1])
    ws.append(["B", 2])

    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=3))
    ws.add_chart(chart, "D2")

    charts = chart_metadata(ws, {})

    assert chart_boxes(ws, {})[0].range_ref.startswith("D2:")
    assert charts[0]["kind"] == "BarChart"
    assert charts[0]["range_ref"].startswith("D2:")
    assert {source["role"] for source in charts[0]["sources"]} == {"cat", "val"}


def test_chart_export_removes_empty_chart_dir(tmp_path):
    wb = Workbook()
    ws = wb.active
    chart_dir = tmp_path / "charts"
    chart_dir.mkdir()

    assert extract_sheet_charts_to_dir(ws, tmp_path) == []
    assert not chart_dir.exists()


def test_run_and_write_skips_snapshots_by_default(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    workbook_path = tmp_path / "region_image.xlsx"
    out_dir = tmp_path / "out"
    wb.save(workbook_path)

    run_and_write(
        workbook_path,
        out_dir=out_dir,
        config_overrides={"include_images": False, "extract_embedded_images": False, "extract_chart_images": False},
    )

    data = json.loads((out_dir / "Sheet1" / "info_regions.json").read_text(encoding="utf-8"))
    assert data["region_images"] == []
    assert not (out_dir / "Sheet1" / "snapshot_plan.json").exists()


def test_run_and_write_samples_large_regions(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Large"
    ws.append(["A", "B", "C"])
    for row in range(2, 151):
        ws.append([row, row * 2, row * 3])
    workbook_path = tmp_path / "large.xlsx"
    out_dir = tmp_path / "out"
    wb.save(workbook_path)

    run_and_write(
        workbook_path,
        out_dir=out_dir,
        config_overrides={"include_images": False, "extract_embedded_images": False, "extract_chart_images": False},
        write_snapshots=True,
    )

    plan = json.loads((out_dir / "Large" / "snapshot_plan.json").read_text(encoding="utf-8"))
    assert plan[0]["strategy"] == "large_table_sampled"
    assert any(snapshot["kind"] == "overview" and snapshot["render"] is False for snapshot in plan[0]["snapshots"])
    rendered = [snapshot for snapshot in plan[0]["snapshots"] if snapshot.get("render", True)]
    assert rendered
    assert all((out_dir / "Large" / snapshot["path"]).exists() for snapshot in rendered)


def test_merge_structure_changes_are_clustered_with_margin():
    wb = Workbook()
    ws = wb.active
    for row in range(1, 21):
        for col in range(1, 9):
            ws.cell(row, col).value = "x"
    for row, ranges in {
        1: ["A1:C1", "D1:F1", "G1:H1"],
        3: ["A3:B3", "C3:E3", "F3:H3"],
        7: ["A7:D7", "E7:H7"],
    }.items():
        for range_ref in ranges:
            ws.merge_cells(range_ref)

    assert [box.range_ref for box in _merge_structure_boxes(ws, Box(1, 1, 20, 8))] == ["A1:H12"]
