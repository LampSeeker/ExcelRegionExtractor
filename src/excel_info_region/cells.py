from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from .components import clip_box_to_bounds
from .schema import Box


def is_non_empty(value: Any) -> bool:
    return value is not None and value != ""


def merged_boxes_with_values(ws: Worksheet) -> list[Box]:
    boxes: list[Box] = []
    for rng in ws.merged_cells.ranges:
        top_left = ws.cell(rng.min_row, rng.min_col)
        if is_non_empty(top_left.value):
            boxes.append(Box(rng.min_row, rng.min_col, rng.max_row, rng.max_col))
    return boxes


def is_hidden_cell(ws: Worksheet, row: int, col: int, config: dict[str, Any]) -> bool:
    if not config.get("respect_hidden_rows_cols", False):
        return False
    return is_hidden_row(ws, row, config) or is_hidden_col(ws, col, config)


def is_hidden_row(ws: Worksheet, row: int, config: dict[str, Any]) -> bool:
    return bool(config.get("respect_hidden_rows_cols", False) and ws.row_dimensions[row].hidden)


def is_hidden_col(ws: Worksheet, col: int, config: dict[str, Any]) -> bool:
    col_letter = get_column_letter(col)
    return bool(config.get("respect_hidden_rows_cols", False) and ws.column_dimensions[col_letter].hidden)


def visible_rows_cols_in_box(ws: Worksheet, box: Box, config: dict[str, Any]) -> tuple[list[int], list[int]]:
    rows = [
        row
        for row in range(box.min_row, box.max_row + 1)
        if not is_hidden_row(ws, row, config)
    ]
    cols = [
        col
        for col in range(box.min_col, box.max_col + 1)
        if not is_hidden_col(ws, col, config)
    ]
    return rows, cols


def visible_box(ws: Worksheet, box: Box, config: dict[str, Any]) -> Box | None:
    rows, cols = visible_rows_cols_in_box(ws, box, config)
    if not rows or not cols:
        return None
    return Box(min(rows), min(cols), max(rows), max(cols))


def print_area_bounds(ws: Worksheet) -> Box | None:
    print_area = ws.print_area
    if not print_area:
        return None

    boxes: list[Box] = []
    for area in str(print_area).split(","):
        ref = area.split("!", 1)[-1].replace("$", "").strip().strip("'")
        if not ref:
            continue
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        boxes.append(Box(min_row, min_col, max_row, max_col))

    if not boxes:
        return None

    bounds = boxes[0]
    for box in boxes[1:]:
        bounds = bounds.union(box)
    return bounds


def apply_print_area_bounds(ws: Worksheet, bounds: Box | None, config: dict[str, Any]) -> Box | None:
    if not config.get("use_print_area_bounds", False):
        return bounds

    print_bounds = print_area_bounds(ws)
    if print_bounds is None:
        return bounds
    if bounds is None:
        return print_bounds
    return clip_box_to_bounds(bounds, print_bounds)


def collect_cell_occupied(ws: Worksheet, bounds: Box | None, config: dict[str, Any]) -> set[tuple[int, int]]:
    if bounds is None:
        return set()
    occupied: set[tuple[int, int]] = set()

    if config.get("include_values", True):
        for (row, col), cell in ws._cells.items():
            if bounds.contains(row, col) and is_non_empty(cell.value) and not is_hidden_cell(ws, row, col, config):
                occupied.add((row, col))

    if config.get("include_merged_cells", True):
        for box in merged_boxes_with_values(ws):
            for row in range(max(bounds.min_row, box.min_row), min(bounds.max_row, box.max_row) + 1):
                for col in range(max(bounds.min_col, box.min_col), min(bounds.max_col, box.max_col) + 1):
                    if not is_hidden_cell(ws, row, col, config):
                        occupied.add((row, col))

    return occupied
