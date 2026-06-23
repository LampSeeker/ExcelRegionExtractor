from __future__ import annotations

from collections import deque
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from .components import (
    dedupe_boxes,
    intersection_area,
    overlap_ratio_on_axis,
    union_find_groups,
)
from .cells import is_hidden_cell
from .schema import Box


def _side_has_style(side) -> bool:
    return side is not None and getattr(side, "style", None) is not None


def cell_has_border(cell, *, strong_only: bool = False) -> bool:
    border = cell.border
    sides = [border.left, border.right, border.top, border.bottom]

    if not strong_only:
        return any(_side_has_style(side) for side in sides)

    strong_styles = {
        "medium",
        "thick",
        "double",
        "mediumDashed",
        "mediumDashDot",
        "mediumDashDotDot",
        "slantDashDot",
    }
    return any(getattr(side, "style", None) in strong_styles for side in sides)


def collect_border_occupied(ws: Worksheet, bounds: Box | None, config: dict[str, Any]) -> set[tuple[int, int]]:
    if bounds is None or not config.get("use_borders", True):
        return set()

    strong_only = bool(config.get("strong_borders_only", False))
    occupied: set[tuple[int, int]] = set()

    for (row, col), cell in ws._cells.items():
        if bounds.contains(row, col) and not is_hidden_cell(ws, row, col, config) and cell_has_border(cell, strong_only=strong_only):
            occupied.add((row, col))

    # Merged cells often store border information only on edge cells.
    if config.get("include_merged_cells", True):
        for rng in ws.merged_cells.ranges:
            box = Box(rng.min_row, rng.min_col, rng.max_row, rng.max_col)
            if not box.intersects(bounds):
                continue

            edge_cells = []
            for col in range(rng.min_col, rng.max_col + 1):
                edge_cells.append(ws.cell(rng.min_row, col))
                edge_cells.append(ws.cell(rng.max_row, col))
            for row in range(rng.min_row, rng.max_row + 1):
                edge_cells.append(ws.cell(row, rng.min_col))
                edge_cells.append(ws.cell(row, rng.max_col))

            if any(cell_has_border(cell, strong_only=strong_only) for cell in edge_cells):
                for row in range(max(bounds.min_row, rng.min_row), min(bounds.max_row, rng.max_row) + 1):
                    for col in range(max(bounds.min_col, rng.min_col), min(bounds.max_col, rng.max_col) + 1):
                        if not is_hidden_cell(ws, row, col, config):
                            occupied.add((row, col))

    return occupied


def should_expand_to_border_shell(value_box: Box, border_box: Box, config: dict[str, Any]) -> bool:
    inter = intersection_area(value_box, border_box)
    if inter <= 0:
        return False

    value_overlap = inter / max(1, value_box.area)
    border_overlap = inter / max(1, border_box.area)

    if value_overlap < float(config.get("border_expand_min_value_overlap", 0.80)):
        return False

    # Border expansion is bbox correction, not section grouping.
    max_area_ratio = float(config.get("border_expand_max_area_ratio", 3.0))
    if border_box.area > value_box.area * max_area_ratio:
        return False

    max_extra_rows = int(config.get("border_expand_max_extra_rows", 3))
    max_extra_cols = int(config.get("border_expand_max_extra_cols", 3))
    extra_top = max(0, value_box.min_row - border_box.min_row)
    extra_bottom = max(0, border_box.max_row - value_box.max_row)
    extra_left = max(0, value_box.min_col - border_box.min_col)
    extra_right = max(0, border_box.max_col - value_box.max_col)

    if max(extra_top, extra_bottom) > max_extra_rows:
        return False
    if max(extra_left, extra_right) > max_extra_cols:
        return False

    # Avoid over-expanding a tiny value region into a large outlined region.
    return border_overlap >= float(config.get("border_expand_min_border_overlap", 0.10))


def expand_cell_boxes_with_borders(
    cell_boxes: list[Box],
    border_boxes: list[Box],
    config: dict[str, Any],
) -> list[Box]:
    if not config.get("use_borders", True) or not border_boxes:
        return cell_boxes

    expanded: list[Box] = []
    for value_box in cell_boxes:
        current = value_box
        for border_box in border_boxes:
            if should_expand_to_border_shell(current, border_box, config):
                current = current.union(border_box)
        expanded.append(current)

    if config.get("add_border_only_regions", False):
        for border_box in border_boxes:
            has_value = any(
                border_box.contains_box(value_box) or intersection_area(border_box, value_box) > 0
                for value_box in cell_boxes
            )
            if not has_value:
                expanded.append(border_box)

    return dedupe_boxes(expanded)


BorderEdge = tuple[str, int, int]  # ("h", row_line, col) or ("v", row, col_line)


def cell_has_border_side(side, *, strong_only: bool = False) -> bool:
    if not _side_has_style(side):
        return False
    if not strong_only:
        return True
    return getattr(side, "style", None) in {
        "medium",
        "thick",
        "double",
        "mediumDashed",
        "mediumDashDot",
        "mediumDashDotDot",
        "slantDashDot",
    }


def collect_border_edges(ws: Worksheet, bounds: Box | None, config: dict[str, Any]) -> set[BorderEdge]:
    if bounds is None or not config.get("use_border_contact_merge", False):
        return set()

    strong_only = bool(config.get("border_contact_strong_only", False))
    edges: set[BorderEdge] = set()

    def add_cell_edges(row: int, col: int, cell) -> None:
        b = cell.border
        if _side_has_style(b.top) and (not strong_only or cell_has_border_side(b.top, strong_only=True)):
            edges.add(("h", row, col))
        if _side_has_style(b.bottom) and (not strong_only or cell_has_border_side(b.bottom, strong_only=True)):
            edges.add(("h", row + 1, col))
        if _side_has_style(b.left) and (not strong_only or cell_has_border_side(b.left, strong_only=True)):
            edges.add(("v", row, col))
        if _side_has_style(b.right) and (not strong_only or cell_has_border_side(b.right, strong_only=True)):
            edges.add(("v", row, col + 1))

    for (row, col), cell in ws._cells.items():
        if bounds.contains(row, col) and not is_hidden_cell(ws, row, col, config):
            add_cell_edges(row, col, cell)

    if config.get("include_merged_cells", True):
        for rng in ws.merged_cells.ranges:
            box = Box(rng.min_row, rng.min_col, rng.max_row, rng.max_col)
            if not box.intersects(bounds):
                continue

            for col in range(rng.min_col, rng.max_col + 1):
                top_cell = ws.cell(rng.min_row, col)
                bottom_cell = ws.cell(rng.max_row, col)
                if not is_hidden_cell(ws, rng.min_row, col, config) and _side_has_style(top_cell.border.top) and (not strong_only or cell_has_border_side(top_cell.border.top, strong_only=True)):
                    edges.add(("h", rng.min_row, col))
                if not is_hidden_cell(ws, rng.max_row, col, config) and _side_has_style(bottom_cell.border.bottom) and (not strong_only or cell_has_border_side(bottom_cell.border.bottom, strong_only=True)):
                    edges.add(("h", rng.max_row + 1, col))

            for row in range(rng.min_row, rng.max_row + 1):
                left_cell = ws.cell(row, rng.min_col)
                right_cell = ws.cell(row, rng.max_col)
                if not is_hidden_cell(ws, row, rng.min_col, config) and _side_has_style(left_cell.border.left) and (not strong_only or cell_has_border_side(left_cell.border.left, strong_only=True)):
                    edges.add(("v", row, rng.min_col))
                if not is_hidden_cell(ws, row, rng.max_col, config) and _side_has_style(right_cell.border.right) and (not strong_only or cell_has_border_side(right_cell.border.right, strong_only=True)):
                    edges.add(("v", row, rng.max_col + 1))

    return edges


def edge_endpoints(edge: BorderEdge) -> tuple[tuple[int, int], tuple[int, int]]:
    kind, a, b = edge
    if kind == "h":
        return (a, b), (a, b + 1)
    return (a, b), (a + 1, b)


def border_edge_components(edges: set[BorderEdge]) -> dict[BorderEdge, int]:
    endpoint_to_edges: dict[tuple[int, int], list[BorderEdge]] = {}
    for edge in edges:
        p1, p2 = edge_endpoints(edge)
        endpoint_to_edges.setdefault(p1, []).append(edge)
        endpoint_to_edges.setdefault(p2, []).append(edge)

    edge_to_component: dict[BorderEdge, int] = {}
    visited: set[BorderEdge] = set()
    component_id = 0

    for start in sorted(edges):
        if start in visited:
            continue
        component_id += 1
        q = deque([start])
        visited.add(start)
        edge_to_component[start] = component_id

        while q:
            edge = q.popleft()
            for point in edge_endpoints(edge):
                for nxt in endpoint_to_edges.get(point, []):
                    if nxt not in visited:
                        visited.add(nxt)
                        edge_to_component[nxt] = component_id
                        q.append(nxt)

    return edge_to_component


def perimeter_edges_by_side(box: Box, *, tolerance: int = 0) -> dict[str, set[BorderEdge]]:
    min_row = max(1, box.min_row - tolerance)
    min_col = max(1, box.min_col - tolerance)
    max_row = box.max_row + tolerance
    max_col = box.max_col + tolerance

    return {
        "top": {("h", min_row, col) for col in range(min_col, max_col + 1)},
        "bottom": {("h", max_row + 1, col) for col in range(min_col, max_col + 1)},
        "left": {("v", row, min_col) for row in range(min_row, max_row + 1)},
        "right": {("v", row, max_col + 1) for row in range(min_row, max_row + 1)},
    }


def touched_border_component_sides(
    box: Box,
    edge_to_component: dict[BorderEdge, int],
    config: dict[str, Any],
) -> dict[int, set[str]]:
    tolerance = int(config.get("border_contact_tolerance_cells", 0))
    min_edges_per_side = int(
        config.get(
            "border_contact_min_edges_per_side",
            config.get("border_contact_min_edges", 1),
        )
    )

    component_side_counts: dict[int, dict[str, int]] = {}
    for side_name, edges in perimeter_edges_by_side(box, tolerance=tolerance).items():
        for edge in edges:
            component_id = edge_to_component.get(edge)
            if component_id is None:
                continue
            component_side_counts.setdefault(component_id, {})
            component_side_counts[component_id][side_name] = (
                component_side_counts[component_id].get(side_name, 0) + 1
            )

    result: dict[int, set[str]] = {}
    for component_id, side_counts in component_side_counts.items():
        sides = {
            side_name
            for side_name, count in side_counts.items()
            if count >= min_edges_per_side
        }
        if sides:
            result[component_id] = sides
    return result


def boxes_are_contact_merge_neighbors(a: Box, b: Box, config: dict[str, Any]) -> bool:
    if a.contains_box(b) or b.contains_box(a):
        return False

    max_gap = int(config.get("border_contact_merge_max_gap", 1))
    min_axis_overlap = float(config.get("border_contact_merge_min_axis_overlap", 0.80))

    if a.max_row < b.min_row:
        row_gap = b.min_row - a.max_row - 1
        return row_gap <= max_gap and overlap_ratio_on_axis(a, b, axis="col") >= min_axis_overlap
    if b.max_row < a.min_row:
        row_gap = a.min_row - b.max_row - 1
        return row_gap <= max_gap and overlap_ratio_on_axis(a, b, axis="col") >= min_axis_overlap

    if a.max_col < b.min_col:
        col_gap = b.min_col - a.max_col - 1
        return col_gap <= max_gap and overlap_ratio_on_axis(a, b, axis="row") >= min_axis_overlap
    if b.max_col < a.min_col:
        col_gap = a.min_col - b.max_col - 1
        return col_gap <= max_gap and overlap_ratio_on_axis(a, b, axis="row") >= min_axis_overlap

    return True


def merge_boxes_by_border_contact(
    cell_boxes: list[Box],
    ws: Worksheet,
    bounds: Box | None,
    config: dict[str, Any],
) -> list[Box]:
    if not config.get("use_border_contact_merge", False) or len(cell_boxes) < 2:
        return cell_boxes

    edges = collect_border_edges(ws, bounds, config)
    if not edges:
        return cell_boxes

    edge_to_component = border_edge_components(edges)
    min_touched_sides = int(config.get("border_contact_min_touched_sides", 2))
    component_to_indices: dict[int, list[int]] = {}

    for idx, box in enumerate(cell_boxes):
        touched_sides_by_component = touched_border_component_sides(box, edge_to_component, config)
        for component_id, sides in touched_sides_by_component.items():
            if len(sides) >= min_touched_sides:
                component_to_indices.setdefault(component_id, []).append(idx)

    candidate_pairs: set[tuple[int, int]] = set()
    for indices in component_to_indices.values():
        unique_indices = sorted(set(indices))
        if len(unique_indices) < 2:
            continue
        for pos, a_idx in enumerate(unique_indices):
            for b_idx in unique_indices[pos + 1:]:
                if boxes_are_contact_merge_neighbors(cell_boxes[a_idx], cell_boxes[b_idx], config):
                    candidate_pairs.add((a_idx, b_idx))

    if not candidate_pairs:
        return cell_boxes

    output: list[Box] = []
    for group in union_find_groups(list(range(len(cell_boxes))), sorted(candidate_pairs)):
        if len(group) == 1:
            output.append(cell_boxes[group[0]])
            continue

        merged = cell_boxes[group[0]]
        for idx in group[1:]:
            merged = merged.union(cell_boxes[idx])

        total_area = sum(cell_boxes[idx].area for idx in group)
        max_area_ratio = float(config.get("border_contact_merge_max_area_ratio", 2.5))
        if merged.area > total_area * max_area_ratio:
            output.extend(cell_boxes[idx] for idx in group)
        else:
            output.append(merged)

    return dedupe_boxes(output)
