from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .borders import (
    collect_border_occupied,
    expand_cell_boxes_with_borders,
    merge_boxes_by_border_contact,
)
from .cells import (
    apply_print_area_bounds,
    collect_cell_occupied,
    is_hidden_cell,
    is_non_empty,
    merged_boxes_with_values,
    visible_box,
)
from .chart_regions import chart_boxes
from .components import clip_box_to_bounds, connected_components_from_cells, remove_exact_or_contained_duplicates
from .image_regions import image_boxes
from .schema import Box


def open_workbook(path: str | Path, *, data_only: bool = False):
    return load_workbook(path, read_only=False, data_only=data_only)


def iter_target_sheets(workbook, sheet_name: str | None = None, config: dict[str, Any] | None = None):
    if sheet_name:
        yield workbook[sheet_name]
    else:
        skip_hidden = bool((config or {}).get("respect_hidden_sheets", False))
        for ws in workbook.worksheets:
            if skip_hidden and ws.sheet_state != "visible":
                continue
            yield ws


def effective_bounds(ws: Worksheet, workbook_path: str | Path | None, config: dict[str, Any]) -> Box | None:
    rows: list[int] = []
    cols: list[int] = []

    for (row, col), cell in ws._cells.items():
        if is_non_empty(cell.value) and not is_hidden_cell(ws, row, col, config):
            rows.append(row)
            cols.append(col)

    for box in merged_boxes_with_values(ws):
        box = visible_box(ws, box, config)
        if box is None:
            continue
        rows.extend([box.min_row, box.max_row])
        cols.extend([box.min_col, box.max_col])

    for box in image_boxes(ws, workbook_path, config):
        box = visible_box(ws, box, config)
        if box is None:
            continue
        rows.extend([box.min_row, box.max_row])
        cols.extend([box.min_col, box.max_col])

    for box in chart_boxes(ws, config):
        box = visible_box(ws, box, config)
        if box is None:
            continue
        rows.extend([box.min_row, box.max_row])
        cols.extend([box.min_col, box.max_col])

    if not rows or not cols:
        return apply_print_area_bounds(ws, None, config)

    pad_r = int(config.get("bounds_padding_rows", 0))
    pad_c = int(config.get("bounds_padding_cols", 0))
    bounds = Box(
        max(1, min(rows) - pad_r),
        max(1, min(cols) - pad_c),
        min(1048576, max(rows) + pad_r),
        min(16384, max(cols) + pad_c),
    )
    return apply_print_area_bounds(ws, bounds, config)


def extract_info_regions_from_sheet(
    ws: Worksheet,
    *,
    workbook_path: str | Path | None = None,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    cfg = config or {}
    bounds = effective_bounds(ws, workbook_path, cfg)
    if bounds is None:
        return {
            "sheet_name": ws.title,
            "info_regions": [],
        }

    cell_boxes = connected_components_from_cells(
        collect_cell_occupied(ws, bounds, cfg),
        connectivity=int(cfg.get("connectivity", 8)),
        min_occupied_cells=int(cfg.get("min_occupied_cells", 1)),
    )

    border_boxes = connected_components_from_cells(
        collect_border_occupied(ws, bounds, cfg),
        connectivity=int(cfg.get("border_connectivity", cfg.get("connectivity", 8))),
        min_occupied_cells=int(cfg.get("min_border_cells", 2)),
    )
    cell_boxes = expand_cell_boxes_with_borders(cell_boxes, border_boxes, cfg)
    cell_boxes = merge_boxes_by_border_contact(cell_boxes, ws, bounds, cfg)

    # Images stay separate so drawings and adjacent tables do not over-merge.
    clipped_image_boxes = [
        clipped
        for box in image_boxes(ws, workbook_path, cfg)
        if (visible := visible_box(ws, box, cfg)) is not None
        if (clipped := clip_box_to_bounds(visible, bounds)) is not None
    ]
    clipped_chart_boxes = [
        clipped
        for box in chart_boxes(ws, cfg)
        if (visible := visible_box(ws, box, cfg)) is not None
        if (clipped := clip_box_to_bounds(visible, bounds)) is not None
    ]
    boxes = remove_exact_or_contained_duplicates(
        [*cell_boxes, *clipped_image_boxes, *clipped_chart_boxes],
        cfg,
    )

    regions = [
        box.range_ref
        for box in sorted(boxes, key=lambda b: (b.min_row, b.min_col, b.max_row, b.max_col))
    ]

    return {
        "sheet_name": ws.title,
        "info_regions": regions,
    }


def extract_workbook_info_regions(
    workbook_path: str | Path,
    *,
    sheet_name: str | None = None,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    cfg = dict(config or {})
    cfg["workbook_path"] = str(workbook_path)

    wb = open_workbook(workbook_path, data_only=False)
    result = {
        "workbook": str(workbook_path),
        "sheets": {},
    }
    for ws in iter_target_sheets(wb, sheet_name, cfg):
        result["sheets"][ws.title] = extract_info_regions_from_sheet(
            ws,
            workbook_path=workbook_path,
            config=cfg,
        )
    return result


def summarize_workbook_result(result: dict[str, Any]) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for sheet_name, data in result["sheets"].items():
        regions = data.get("regions", data.get("info_regions", []))
        rows.append({
            "sheet_name": sheet_name,
            "region_count": len(regions),
            "regions": regions,
            "image_count": len(data.get("images", [])),
            "images": data.get("images", []),
            "chart_count": len(data.get("charts", [])),
            "charts": data.get("charts", []),
        })
    return {
        "workbook": result["workbook"],
        "sheets": rows,
    }
