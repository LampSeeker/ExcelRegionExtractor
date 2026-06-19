
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any
import posixpath
import zipfile
import xml.etree.ElementTree as ET

from PIL import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .schema import Box


EMU_PER_PIXEL = 9525

NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
}


@dataclass(frozen=True)
class Marker:
    row: int  # one-based
    col: int  # one-based
    row_off: int = 0
    col_off: int = 0


@dataclass(frozen=True)
class DrawingImage:
    name: str
    media_path: str
    data: bytes
    from_marker: Marker
    to_marker: Marker | None = None
    ext_cx: int | None = None
    ext_cy: int | None = None
    child_rel_x: float | None = None
    child_rel_y: float | None = None
    child_rel_w: float | None = None
    child_rel_h: float | None = None
    is_group_child: bool = False

    @property
    def is_grouped(self) -> bool:
        return self.is_group_child


def _resolve_target(base_dir: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(base_dir, target))


def _read_rels(zf: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
    if rels_path not in zf.namelist():
        return {}
    root = ET.fromstring(zf.read(rels_path))
    out: dict[str, str] = {}
    base_dir = posixpath.dirname(rels_path.replace("_rels/", ""))
    if base_dir.endswith(".xml"):
        base_dir = posixpath.dirname(base_dir)
    # For xl/worksheets/_rels/sheet1.xml.rels -> base dir should be xl/worksheets.
    if "/_rels/" in rels_path:
        left, right = rels_path.split("/_rels/", 1)
        base_dir = left
    for rel in root:
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rid and target:
            out[rid] = _resolve_target(base_dir, target)
    return out


def _workbook_sheet_path(zf: zipfile.ZipFile, sheet_name: str) -> str | None:
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = _read_rels(zf, "xl/_rels/workbook.xml.rels")
    for sheet in wb_root.findall(".//main:sheet", NS):
        if sheet.attrib.get("name") != sheet_name:
            continue
        rid = sheet.attrib.get(f"{{{NS['r']}}}id")
        if rid and rid in rels:
            return rels[rid]
    return None


def _sheet_drawing_path(zf: zipfile.ZipFile, sheet_path: str) -> str | None:
    if sheet_path not in zf.namelist():
        return None
    root = ET.fromstring(zf.read(sheet_path))
    drawing = root.find("main:drawing", NS)
    if drawing is None:
        return None
    rid = drawing.attrib.get(f"{{{NS['r']}}}id")
    if not rid:
        return None
    dirname = posixpath.dirname(sheet_path)
    basename = posixpath.basename(sheet_path)
    rels_path = posixpath.join(dirname, "_rels", f"{basename}.rels")
    rels = _read_rels(zf, rels_path)
    return rels.get(rid)


def _drawing_rels_path(drawing_path: str) -> str:
    dirname = posixpath.dirname(drawing_path)
    basename = posixpath.basename(drawing_path)
    return posixpath.join(dirname, "_rels", f"{basename}.rels")


def _marker(el: ET.Element | None) -> Marker | None:
    if el is None:
        return None
    def val(name: str, default: int = 0) -> int:
        node = el.find(f"xdr:{name}", NS)
        if node is None or node.text is None:
            return default
        return int(float(node.text))
    # xdr col/row are zero-based.
    return Marker(
        row=val("row") + 1,
        col=val("col") + 1,
        row_off=val("rowOff"),
        col_off=val("colOff"),
    )


def _int_attr(el: ET.Element | None, key: str, default: int = 0) -> int:
    if el is None:
        return default
    try:
        return int(float(el.attrib.get(key, default)))
    except Exception:
        return default


def _shape_name(pic: ET.Element) -> str:
    c_nv = pic.find(".//xdr:cNvPr", NS)
    return c_nv.attrib.get("name", "") if c_nv is not None else ""


def _pic_rel_id(pic: ET.Element) -> str | None:
    blip = pic.find(".//a:blip", NS)
    if blip is None:
        return None
    return blip.attrib.get(f"{{{NS['r']}}}embed")


def _pic_xfrm(pic: ET.Element) -> tuple[int, int, int, int] | None:
    xfrm = pic.find(".//xdr:spPr/a:xfrm", NS)
    if xfrm is None:
        return None
    off = xfrm.find("a:off", NS)
    ext = xfrm.find("a:ext", NS)
    return (
        _int_attr(off, "x"),
        _int_attr(off, "y"),
        _int_attr(ext, "cx"),
        _int_attr(ext, "cy"),
    )


def _group_transform(grp: ET.Element) -> tuple[int, int, int, int, int, int, int, int] | None:
    xfrm = grp.find("xdr:grpSpPr/a:xfrm", NS)
    if xfrm is None:
        return None
    off = xfrm.find("a:off", NS)
    ext = xfrm.find("a:ext", NS)
    ch_off = xfrm.find("a:chOff", NS)
    ch_ext = xfrm.find("a:chExt", NS)
    return (
        _int_attr(off, "x"),
        _int_attr(off, "y"),
        _int_attr(ext, "cx"),
        _int_attr(ext, "cy"),
        _int_attr(ch_off, "x"),
        _int_attr(ch_off, "y"),
        max(1, _int_attr(ch_ext, "cx", 1)),
        max(1, _int_attr(ch_ext, "cy", 1)),
    )


def _image_size_emu(data: bytes) -> tuple[int, int]:
    try:
        with Image.open(BytesIO(data)) as im:
            return int(im.width * EMU_PER_PIXEL), int(im.height * EMU_PER_PIXEL)
    except Exception:
        return 64 * EMU_PER_PIXEL, 64 * EMU_PER_PIXEL


def extract_drawing_images(workbook_path: str | Path, sheet_name: str) -> list[DrawingImage]:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        return []

    with zipfile.ZipFile(workbook_path) as zf:
        sheet_path = _workbook_sheet_path(zf, sheet_name)
        if not sheet_path:
            return []
        drawing_path = _sheet_drawing_path(zf, sheet_path)
        if not drawing_path or drawing_path not in zf.namelist():
            return []

        rels = _read_rels(zf, _drawing_rels_path(drawing_path))
        root = ET.fromstring(zf.read(drawing_path))
        images: list[DrawingImage] = []

        for anchor in list(root):
            tag = anchor.tag.split("}")[-1]
            if tag not in {"oneCellAnchor", "twoCellAnchor"}:
                continue
            start = _marker(anchor.find("xdr:from", NS))
            end = _marker(anchor.find("xdr:to", NS))
            if start is None:
                continue
            anchor_ext = anchor.find("xdr:ext", NS)
            ext_cx = _int_attr(anchor_ext, "cx") if anchor_ext is not None else None
            ext_cy = _int_attr(anchor_ext, "cy") if anchor_ext is not None else None

            # Direct picture anchors.
            for pic in anchor.findall("xdr:pic", NS):
                rid = _pic_rel_id(pic)
                if not rid or rid not in rels:
                    continue
                media_path = rels[rid]
                if media_path not in zf.namelist():
                    continue
                data = zf.read(media_path)
                cx, cy = _image_size_emu(data)
                images.append(
                    DrawingImage(
                        name=_shape_name(pic),
                        media_path=media_path,
                        data=data,
                        from_marker=start,
                        to_marker=end,
                        ext_cx=ext_cx or cx,
                        ext_cy=ext_cy or cy,
                        is_group_child=False,
                    )
                )

            # Grouped pictures: openpyxl commonly misses these children.
            for grp in anchor.findall("xdr:grpSp", NS):
                gt = _group_transform(grp)
                if gt is None:
                    continue
                _off_x, _off_y, _ext_x, _ext_y, ch_off_x, ch_off_y, ch_ext_x, ch_ext_y = gt
                for pic in grp.findall(".//xdr:pic", NS):
                    rid = _pic_rel_id(pic)
                    xfrm = _pic_xfrm(pic)
                    if not rid or rid not in rels or xfrm is None:
                        continue
                    media_path = rels[rid]
                    if media_path not in zf.namelist():
                        continue
                    pic_off_x, pic_off_y, pic_ext_x, pic_ext_y = xfrm
                    images.append(
                        DrawingImage(
                            name=_shape_name(pic),
                            media_path=media_path,
                            data=zf.read(media_path),
                            from_marker=start,
                            to_marker=end,
                            ext_cx=ext_cx,
                            ext_cy=ext_cy,
                            child_rel_x=(pic_off_x - ch_off_x) / ch_ext_x,
                            child_rel_y=(pic_off_y - ch_off_y) / ch_ext_y,
                            child_rel_w=pic_ext_x / ch_ext_x,
                            child_rel_h=pic_ext_y / ch_ext_y,
                            is_group_child=True,
                        )
                    )

        return images


def _column_width_to_pixels(width: float | None) -> int:
    width = 8.43 if width is None else float(width)
    return max(4, int(width * 7 + 5))


def _row_height_to_pixels(height_pt: float | None) -> int:
    height_pt = 15.0 if height_pt is None else float(height_pt)
    return max(4, int(height_pt * 96 / 72))


def sheet_pixel_axes(ws: Worksheet, max_row: int, max_col: int) -> tuple[dict[int, int], dict[int, int], dict[int, int], dict[int, int]]:
    col_x: dict[int, int] = {}
    col_w: dict[int, int] = {}
    x = 0
    for col in range(1, max_col + 2):
        col_x[col] = x
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width if letter in ws.column_dimensions else None
        col_w[col] = _column_width_to_pixels(width)
        x += col_w[col]

    row_y: dict[int, int] = {}
    row_h: dict[int, int] = {}
    y = 0
    for row in range(1, max_row + 2):
        row_y[row] = y
        height = ws.row_dimensions[row].height if row in ws.row_dimensions else None
        row_h[row] = _row_height_to_pixels(height)
        y += row_h[row]

    return col_x, row_y, col_w, row_h


def marker_to_px(marker: Marker, col_x: dict[int, int], row_y: dict[int, int]) -> tuple[float, float]:
    return (
        col_x.get(marker.col, 0) + marker.col_off / EMU_PER_PIXEL,
        row_y.get(marker.row, 0) + marker.row_off / EMU_PER_PIXEL,
    )


def drawing_image_pixel_box(img: DrawingImage, ws: Worksheet, col_x: dict[int, int], row_y: dict[int, int]) -> tuple[float, float, float, float]:
    x0, y0 = marker_to_px(img.from_marker, col_x, row_y)
    if img.to_marker is not None:
        x1, y1 = marker_to_px(img.to_marker, col_x, row_y)
    else:
        x1 = x0 + (img.ext_cx or 1) / EMU_PER_PIXEL
        y1 = y0 + (img.ext_cy or 1) / EMU_PER_PIXEL

    w = max(1.0, x1 - x0)
    h = max(1.0, y1 - y0)

    if img.is_group_child and img.child_rel_x is not None:
        cx0 = x0 + img.child_rel_x * w
        cy0 = y0 + (img.child_rel_y or 0.0) * h
        cx1 = cx0 + (img.child_rel_w or 0.01) * w
        cy1 = cy0 + (img.child_rel_h or 0.01) * h
        return cx0, cy0, cx1, cy1
    return x0, y0, x1, y1


def pixel_box_to_cell_box(px_box: tuple[float, float, float, float], col_x: dict[int, int], row_y: dict[int, int], max_row: int, max_col: int) -> Box:
    x0, y0, x1, y1 = px_box

    def col_at(x: float) -> int:
        last = 1
        for col in range(1, max_col + 1):
            if col_x.get(col, 0) <= x:
                last = col
            else:
                break
        return last

    def row_at(y: float) -> int:
        last = 1
        for row in range(1, max_row + 1):
            if row_y.get(row, 0) <= y:
                last = row
            else:
                break
        return last

    min_col = max(1, col_at(x0))
    max_col_v = max(min_col, min(max_col, col_at(x1)))
    min_row = max(1, row_at(y0))
    max_row_v = max(min_row, min(max_row, row_at(y1)))
    return Box(min_row, min_col, max_row_v, max_col_v)


def drawing_image_boxes(workbook_path: str | Path, sheet_name: str, ws: Worksheet) -> list[Box]:
    imgs = extract_drawing_images(workbook_path, sheet_name)
    if not imgs:
        return []
    max_row = max(ws.max_row, max((i.to_marker.row if i.to_marker else i.from_marker.row + 50) for i in imgs))
    max_col = max(ws.max_column, max((i.to_marker.col if i.to_marker else i.from_marker.col + 30) for i in imgs))
    col_x, row_y, _col_w, _row_h = sheet_pixel_axes(ws, max_row + 2, max_col + 2)
    boxes: list[Box] = []
    for img in imgs:
        box = pixel_box_to_cell_box(drawing_image_pixel_box(img, ws, col_x, row_y), col_x, row_y, max_row + 2, max_col + 2)
        boxes.append(box)
    return boxes
