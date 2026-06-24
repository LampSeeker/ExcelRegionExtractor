from __future__ import annotations

from pathlib import Path
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any

from .config import load_config
from .borders import collect_border_edges
from .components import connected_components_from_cells, dedupe_boxes
from .extractor import extract_workbook_info_regions, summarize_workbook_result, open_workbook
from .chart_export import extract_sheet_charts_to_dir
from .io import ensure_dir, safe_name, write_json
from .image_export import extract_sheet_images_to_dir
from .schema import Box
from .visualize import render_region_overlay


def _overlay_regions_from_ranges(ranges: list[str], sheet_name: str) -> list[dict]:
    regions: list[dict] = []
    for idx, range_ref in enumerate(ranges, 1):
        regions.append(_overlay_region(range_ref, sheet_name, f"R{idx:03d}"))
    return regions


def _overlay_region(range_ref: str, sheet_name: str, region_id: str) -> dict:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    return {
        "id": region_id,
        "sheet_name": sheet_name,
        "range_ref": range_ref,
        "min_row": min_row,
        "min_col": min_col,
        "max_row": max_row,
        "max_col": max_col,
        "height": max_row - min_row + 1,
        "width": max_col - min_col + 1,
        "area": (max_row - min_row + 1) * (max_col - min_col + 1),
    }


def _overlay_regions_from_tree(region_tree: list[dict[str, Any]], sheet_name: str) -> list[dict]:
    regions: list[dict] = []
    for idx, root in enumerate(region_tree, 1):
        root_id = f"R{idx:03d}"
        regions.append(_overlay_region(root["range_ref"], sheet_name, root_id))
        for child_idx, child in enumerate(root.get("children", []), 1):
            regions.append(_overlay_region(child["range_ref"], sheet_name, f"{root_id}.C{child_idx:02d}"))
    return regions


def _range_dict(box: Box) -> dict[str, Any]:
    return {
        "range_ref": box.range_ref,
        "min_row": box.min_row,
        "min_col": box.min_col,
        "max_row": box.max_row,
        "max_col": box.max_col,
    }


def _row_signature(ws: Worksheet, row: int, box: Box) -> tuple[bool, ...]:
    return tuple(ws.cell(row, col).value is not None for col in range(box.min_col, box.max_col + 1))


def _change_point_row(ws: Worksheet, box: Box) -> int:
    start = min(box.max_row, box.min_row + 20)
    previous = _row_signature(ws, start, box)
    for row in range(start + 1, box.max_row + 1):
        current = _row_signature(ws, row, box)
        if current != previous:
            return row
        previous = current
    return box.min_row + box.height // 2


def _merge_signature(ws: Worksheet, row: int, box: Box) -> tuple[tuple[int, int], ...]:
    spans: list[tuple[int, int]] = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.max_col >= box.min_col and rng.min_col <= box.max_col:
            min_col = max(box.min_col, rng.min_col)
            max_col = min(box.max_col, rng.max_col)
            if max_col > min_col:
                spans.append((min_col, max_col))
    return tuple(sorted(spans))


def _merge_structure_boxes(ws: Worksheet, box: Box, margin: int = 5) -> list[Box]:
    changes: list[int] = []
    previous = _merge_signature(ws, box.min_row, box)
    if previous:
        changes.append(box.min_row)
    for row in range(box.min_row + 1, box.max_row + 1):
        current = _merge_signature(ws, row, box)
        if current != previous and current:
            changes.append(row)
        previous = current

    clusters: list[list[int]] = []
    for row in changes:
        if clusters and row - clusters[-1][-1] <= margin:
            clusters[-1].append(row)
        else:
            clusters.append([row])

    return [
        Box(max(box.min_row, cluster[0] - margin), box.min_col, min(box.max_row, cluster[-1] + margin), box.max_col)
        for cluster in clusters
    ]


def _merge_snapshot_samples(samples: list[tuple[str, Box, str]], margin: int = 5) -> list[tuple[str, Box, str]]:
    merged: list[tuple[str, Box, str]] = []
    for kind, box, purpose in sorted(samples, key=lambda item: item[1].min_row):
        if merged and box.min_row - merged[-1][1].max_row <= margin:
            prev_kind, prev_box, prev_purpose = merged[-1]
            merged[-1] = (
                prev_kind if kind in prev_kind.split("+") else f"{prev_kind}+{kind}",
                prev_box.union(box),
                prev_purpose if purpose in prev_purpose.split("; ") else f"{prev_purpose}; {purpose}",
            )
        else:
            merged.append((kind, box, purpose))
    return merged


def _region_snapshot_plan(ws: Worksheet, region: dict[str, Any], image_ranges: set[str]) -> dict[str, Any]:
    box = _box_from_range(str(region["range_ref"]))
    merged_count = sum(1 for rng in ws.merged_cells.ranges if box.intersects(Box(rng.min_row, rng.min_col, rng.max_row, rng.max_col)))
    formula_count = sum(
        1
        for row in range(box.min_row, box.max_row + 1)
        for col in range(box.min_col, box.max_col + 1)
        if isinstance(ws.cell(row, col).value, str) and ws.cell(row, col).value.startswith("=")
    )
    if box.range_ref in image_ranges:
        strategy = "embedded_image"
        snapshots: list[dict[str, Any]] = []
    elif box.height > 120 or box.area > 1000:
        strategy = "large_table_sampled"
        change_row = max(box.min_row, min(box.max_row, _change_point_row(ws, box)))
        samples = [("overview", box, "metadata only; do not render the full region")]
        render_samples = [
            ("header", Box(box.min_row, box.min_col, min(box.max_row, box.min_row + 19), box.max_col), "title, column headers, units, notes"),
            ("first_rows", Box(min(box.max_row, box.min_row + 20), box.min_col, min(box.max_row, box.min_row + 59), box.max_col), "early data row pattern"),
            ("change_points", Box(max(box.min_row, change_row - 20), box.min_col, min(box.max_row, change_row + 20), box.max_col), "possible structure change around the sampled row"),
            ("last_rows", Box(max(box.min_row, box.max_row - 47), box.min_col, box.max_row, box.max_col), "ending rows, totals, footnotes"),
        ]
        render_samples.extend(
            ("merge_structure", merge_box, "column merge structure changes")
            for merge_box in _merge_structure_boxes(ws, box)
        )
        samples.extend(_merge_snapshot_samples(render_samples))
        snapshots = [
            {
                "snapshot_id": f"S{idx:03d}",
                "kind": kind,
                "range_ref": sample_box.range_ref,
                "purpose": purpose,
                "render": kind != "overview",
            }
            for idx, (kind, sample_box, purpose) in enumerate(samples, 1)
        ]
    else:
        strategy = "full_region"
        snapshots = [{
            "snapshot_id": "S001",
            "kind": "full_region",
            "range_ref": box.range_ref,
            "purpose": "complete small information region",
            "render": True,
        }]

    return {
        "sheet_name": ws.title,
        "region_id": str(region["id"]),
        "region_range": box.range_ref,
        "row_count": box.height,
        "col_count": box.width,
        "merged_cell_count": merged_count,
        "formula_count": formula_count,
        "strategy": strategy,
        "snapshots": snapshots,
    }


def _snapshot_plan(ws: Worksheet, regions: list[dict[str, Any]], images: list[dict[str, Any]]) -> list[dict[str, Any]]:
    image_ranges = {str(image.get("range_ref")) for image in images}
    return [_region_snapshot_plan(ws, region, image_ranges) for region in regions]


def _render_snapshots(
    ws: Worksheet,
    plan: list[dict[str, Any]],
    sheet_dir: Path,
    viz_cfg: dict[str, Any],
    workbook_path: str | Path,
) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    rel_dir = "snapshots"
    snapshot_dir = sheet_dir / rel_dir
    if snapshot_dir.exists():
        for path in snapshot_dir.glob("*.png"):
            path.unlink()
    for region_plan in plan:
        region_id = str(region_plan["region_id"])
        for snapshot in region_plan["snapshots"]:
            if not snapshot.get("render", True):
                continue
            box = _box_from_range(str(snapshot["range_ref"]))
            bounds = _range_dict(box)
            filename = safe_name(f"{region_id}_{snapshot['snapshot_id']}_{snapshot['kind']}_{box.range_ref}.png")
            render_region_overlay(
                ws,
                [],
                sheet_dir / rel_dir / filename,
                title=f"{ws.title} - {region_id} {snapshot['snapshot_id']} {box.range_ref}",
                bounds=bounds,
                max_rows=int(viz_cfg.get("snapshot_max_rows", viz_cfg.get("max_rows", 160))),
                max_cols=int(viz_cfg.get("snapshot_max_cols", viz_cfg.get("max_cols", 90))),
                pad_rows=int(viz_cfg.get("snapshot_pad_rows", 0)),
                pad_cols=int(viz_cfg.get("snapshot_pad_cols", 0)),
                preserve_dimensions=bool(viz_cfg.get("preserve_dimensions", True)),
                include_images=False,
                include_cell_text=bool(viz_cfg.get("include_cell_text", True)),
                include_merged_cells=bool(viz_cfg.get("include_merged_cells", True)),
                scale=float(viz_cfg.get("scale", 1.0)),
                font_path=viz_cfg.get("font_path"),
                workbook_path=str(workbook_path),
            )
            path = f"{rel_dir}/{filename}"
            snapshot["path"] = path
            output.append({
                "region_id": region_id,
                "snapshot_id": str(snapshot["snapshot_id"]),
                "kind": str(snapshot["kind"]),
                "range_ref": box.range_ref,
                "path": path,
            })
    return output


def _box_from_range(range_ref: str) -> Box:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    return Box(min_row, min_col, max_row, max_col)


def _root_ranges(ranges: list[str]) -> list[str]:
    boxes = [(range_ref, _box_from_range(range_ref)) for range_ref in ranges]
    return [
        range_ref
        for range_ref, box in boxes
        if not any(other_ref != range_ref and other_box.contains_box(box) for other_ref, other_box in boxes)
    ]


def _has_closed_border(ws: Worksheet, box: Box) -> bool:
    edges = collect_border_edges(ws, box, {"use_border_contact_merge": True, "include_merged_cells": True})
    if not edges:
        return False
    return all(("h", box.min_row, col) in edges and ("h", box.max_row + 1, col) in edges for col in range(box.min_col, box.max_col + 1)) and all(
        ("v", row, box.min_col) in edges and ("v", row, box.max_col + 1) in edges for row in range(box.min_row, box.max_row + 1)
    )


def _inner_table_boxes(ws: Worksheet, root: Box) -> list[Box]:
    edges = collect_border_edges(ws, root, {"use_border_contact_merge": True, "include_merged_cells": True})
    if not edges:
        return []

    def has_grid_cell(row: int, col: int) -> bool:
        border = ws.cell(row, col).border
        has_horizontal = bool(getattr(border.top, "style", None) or getattr(border.bottom, "style", None))
        has_vertical = bool(getattr(border.left, "style", None) or getattr(border.right, "style", None))
        return has_horizontal and has_vertical

    def closed(box: Box) -> bool:
        return _has_closed_border(ws, box)

    def isolated(box: Box) -> bool:
        return (
            box.width >= 2
            and box.height >= 2
            and box.min_row > root.min_row
            and box.min_col > root.min_col
            and box.max_row < root.max_row
            and box.max_col < root.max_col
        )

    occupied = {
        (row, col)
        for row in range(root.min_row + 1, root.max_row + 1)
        for col in range(root.min_col, root.max_col + 1)
        if has_grid_cell(row, col)
    }
    candidates: list[Box] = []
    for box in connected_components_from_cells(occupied, connectivity=4, min_occupied_cells=4):
        for candidate in (
            Box(box.min_row, box.min_col, min(root.max_row, box.max_row + 1), box.max_col),
            Box(box.min_row, box.min_col, box.max_row, min(root.max_col, box.max_col + 1)),
        ):
            if candidate != box and closed(candidate):
                box = candidate
        if root.contains_box(box) and isolated(box) and closed(box):
            candidates.append(box)

    kept: list[Box] = []
    for box in sorted(dedupe_boxes(candidates), key=lambda b: b.area, reverse=True):
        if not any(other.contains_box(box) for other in kept):
            kept.append(box)
    return kept


def _region_tree(ranges: list[str], ws: Worksheet | None = None) -> list[dict[str, Any]]:
    base_roots = _root_ranges(ranges)
    roots = [(range_ref, _box_from_range(range_ref)) for range_ref in base_roots]
    children_by_root: dict[str, list[Box]] = {range_ref: [] for range_ref, _ in roots}

    for root_ref, root_box in roots:
        topology_children = _inner_table_boxes(ws, root_box) if ws is not None else []
        children_by_root[root_ref].extend(topology_children)

    return [
        {
            "range_ref": root_ref,
            "children": [
                {"range_ref": child.range_ref}
                for child in sorted(_keep_outer_boxes(children_by_root[root_ref]), key=lambda box: (box.min_row, box.min_col))
            ],
        }
        for root_ref, _ in roots
    ]


def _keep_outer_boxes(boxes: list[Box]) -> list[Box]:
    deduped = dedupe_boxes(boxes)
    return [
        box
        for box in deduped
        if not any(other != box and other.contains_box(box) for other in deduped)
    ]


def run_and_write(
    workbook_path: str | Path,
    *,
    out_dir: str | Path,
    sheet_name: str | None = None,
    config_path: str | Path | None = None,
    config_overrides: dict[str, Any] | None = None,
    write_images: bool = True,
    write_snapshots: bool = False,
) -> dict[str, Any]:
    config = load_config(config_path)
    if config_overrides:
        config.update(config_overrides)
    result = extract_workbook_info_regions(workbook_path, sheet_name=sheet_name, config=config)

    out = ensure_dir(out_dir)

    # Use a normal workbook for extracting embedded images because image anchors live
    # in the drawing layer, not in cached cell values.
    wb_drawings = open_workbook(workbook_path, data_only=False)

    # data_only=True is used only for debug overlay PNGs so formula cache values can
    # be displayed when Excel has saved them.
    wb_values = None
    if write_images and config.get("visualization", {}).get("enabled", True):
        wb_values = open_workbook(workbook_path, data_only=True)

    for sheet, data in result["sheets"].items():
        sheet_dir = ensure_dir(out / safe_name(sheet))

        if config.get("extract_embedded_images", True):
            ws_drawing = wb_drawings[sheet]
            images = extract_sheet_images_to_dir(
                workbook_path,
                ws_drawing,
                sheet_dir,
                rel_dir=str(config.get("embedded_image_dir", "images")),
                config=config,
            )
        else:
            images = []

        charts = extract_sheet_charts_to_dir(
            wb_drawings[sheet],
            sheet_dir,
            rel_dir=str(config.get("chart_image_dir", "charts")),
            config=config,
        )

        # Final user-facing schema:
        # {
        #   "sheet_name": "...",
        #   "regions": ["A1:P1", ...],
        #   "images": [{"name": "...", "range_ref": "...", "path": "..."}]
        # }
        root_regions = _root_ranges(data.get("info_regions", []))
        sheet_output = {
            "sheet_name": data["sheet_name"],
            "regions": root_regions,
            "region_tree": _region_tree(data.get("info_regions", []), wb_drawings[sheet]),
            "region_images": [],
            "snapshot_plan": [],
            "snapshots": [],
            "images": images,
            "charts": charts,
        }

        result["sheets"][sheet] = sheet_output

        if wb_values is not None and sheet_output.get("regions"):
            ws = wb_values[sheet]
            viz_cfg = config.get("visualization", {})
            overlay_regions = _overlay_regions_from_tree(sheet_output["region_tree"], sheet)
            render_region_overlay(
                ws,
                overlay_regions,
                sheet_dir / "info_regions.png",
                title=f"{sheet} - info_regions",
                bounds=None,
                max_rows=int(viz_cfg.get("max_rows", 160)),
                max_cols=int(viz_cfg.get("max_cols", 90)),
                pad_rows=int(viz_cfg.get("pad_rows", 1)),
                pad_cols=int(viz_cfg.get("pad_cols", 1)),
                preserve_dimensions=bool(viz_cfg.get("preserve_dimensions", True)),
                include_images=bool(viz_cfg.get("include_images", True)),
                include_cell_text=bool(viz_cfg.get("include_cell_text", True)),
                include_merged_cells=bool(viz_cfg.get("include_merged_cells", True)),
                scale=float(viz_cfg.get("scale", 1.0)),
                font_path=viz_cfg.get("font_path"),
                workbook_path=str(workbook_path),
            )
            if write_snapshots:
                sheet_output["snapshot_plan"] = _snapshot_plan(ws, overlay_regions, images)
                sheet_output["snapshots"] = _render_snapshots(ws, sheet_output["snapshot_plan"], sheet_dir, viz_cfg, workbook_path)
                sheet_output["region_images"] = sheet_output["snapshots"]

        if write_snapshots:
            write_json(sheet_dir / "snapshot_plan.json", sheet_output["snapshot_plan"])
        write_json(sheet_dir / "info_regions.json", sheet_output)

    summary = summarize_workbook_result(result)
    write_json(out / "info_regions_full.json", result)
    write_json(out / "info_regions_summary.json", summary)

    return result
