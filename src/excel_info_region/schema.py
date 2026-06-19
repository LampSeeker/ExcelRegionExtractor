from __future__ import annotations

from dataclasses import dataclass, asdict
from typing import Any
from openpyxl.utils import get_column_letter


def col_letter(col: int) -> str:
    return get_column_letter(col)


def range_ref(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    return f"{col_letter(min_col)}{min_row}:{col_letter(max_col)}{max_row}"


@dataclass(frozen=True)
class Box:
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def height(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def width(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def area(self) -> int:
        return self.height * self.width

    @property
    def range_ref(self) -> str:
        return range_ref(self.min_row, self.min_col, self.max_row, self.max_col)

    def contains(self, row: int, col: int) -> bool:
        return self.min_row <= row <= self.max_row and self.min_col <= col <= self.max_col

    def contains_box(self, other: "Box") -> bool:
        return (
            self.min_row <= other.min_row
            and self.min_col <= other.min_col
            and self.max_row >= other.max_row
            and self.max_col >= other.max_col
        )

    def intersects(self, other: "Box") -> bool:
        return not (
            self.max_row < other.min_row
            or other.max_row < self.min_row
            or self.max_col < other.min_col
            or other.max_col < self.min_col
        )

    def union(self, other: "Box") -> "Box":
        return Box(
            min(self.min_row, other.min_row),
            min(self.min_col, other.min_col),
            max(self.max_row, other.max_row),
            max(self.max_col, other.max_col),
        )

    def to_dict(self) -> dict[str, Any]:
        return {
            "min_row": self.min_row,
            "min_col": self.min_col,
            "max_row": self.max_row,
            "max_col": self.max_col,
            "range_ref": self.range_ref,
            "height": self.height,
            "width": self.width,
            "area": self.area,
        }


@dataclass
class InfoRegion:
    id: str
    sheet_name: str
    box: Box

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "sheet_name": self.sheet_name,
            **self.box.to_dict(),
        }
