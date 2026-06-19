from __future__ import annotations

from dataclasses import dataclass, field, asdict
from typing import Any
from openpyxl.utils import get_column_letter


def col_letter(col: int) -> str:
    return get_column_letter(col)


def range_ref(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    return f"{col_letter(min_col)}{min_row}:{col_letter(max_col)}{max_row}"


@dataclass(frozen=True)
class Box:
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def height(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def width(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def area(self) -> int:
        return self.height * self.width

    @property
    def range_ref(self) -> str:
        return range_ref(self.min_row, self.min_col, self.max_row, self.max_col)

    def contains(self, row: int, col: int) -> bool:
        return self.min_row <= row <= self.max_row and self.min_col <= col <= self.max_col

    def union(self, other: "Box") -> "Box":
        return Box(
            min(self.min_row, other.min_row),
            min(self.min_col, other.min_col),
            max(self.max_row, other.max_row),
            max(self.max_col, other.max_col),
        )

    def intersects(self, other: "Box") -> bool:
        return not (
            self.max_row < other.min_row
            or other.max_row < self.min_row
            or self.max_col < other.min_col
            or other.max_col < self.min_col
        )

    def row_overlap(self, other: "Box") -> int:
        return max(0, min(self.max_row, other.max_row) - max(self.min_row, other.min_row) + 1)

    def col_overlap(self, other: "Box") -> int:
        return max(0, min(self.max_col, other.max_col) - max(self.min_col, other.min_col) + 1)

    def row_gap(self, other: "Box") -> int:
        if self.row_overlap(other) > 0:
            return 0
        return max(other.min_row - self.max_row - 1, self.min_row - other.max_row - 1, 0)

    def col_gap(self, other: "Box") -> int:
        if self.col_overlap(other) > 0:
            return 0
        return max(other.min_col - self.max_col - 1, self.min_col - other.max_col - 1, 0)

    def to_dict(self) -> dict[str, Any]:
        return {
            "min_row": self.min_row,
            "min_col": self.min_col,
            "max_row": self.max_row,
            "max_col": self.max_col,
            "range_ref": self.range_ref,
            "height": self.height,
            "width": self.width,
            "area": self.area,
        }


@dataclass
class CellSignal:
    row: int
    col: int
    value: str | int | float | bool | None = None
    has_value: bool = False
    is_numeric: bool = False
    is_formula: bool = False
    is_merged: bool = False
    has_image: bool = False
    bold: bool = False
    fill: str | None = None
    alignment: str | None = None
    border_score: float = 0.0
    strong_border_score: float = 0.0

    @property
    def coord(self) -> tuple[int, int]:
        return (self.row, self.col)

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["cell_ref"] = f"{col_letter(self.col)}{self.row}"
        return data


@dataclass
class Region:
    id: str
    sheet_name: str
    box: Box
    algorithm: str
    features: dict[str, Any] = field(default_factory=dict)
    members: list[str] = field(default_factory=list)
    score: float | None = None

    def to_dict(self) -> dict[str, Any]:
        data = {
            "id": self.id,
            "sheet_name": self.sheet_name,
            "algorithm": self.algorithm,
            **self.box.to_dict(),
            "features": self.features,
            "members": self.members,
        }
        if self.score is not None:
            data["score"] = self.score
        return data


@dataclass
class RegionEdge:
    source: str
    target: str
    score: float
    components: dict[str, float] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)
