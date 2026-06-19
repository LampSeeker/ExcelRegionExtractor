from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .schema import Box, CellSignal
from .raw_drawing import drawing_image_boxes


@dataclass
class SheetSignals:
    sheet_name: str
    bounds: Box | None
    cells: dict[tuple[int, int], CellSignal]
    occupied: set[tuple[int, int]]
    merged_ranges: list[Box]
    image_ranges: list[Box]

    def occupied_in_box(self, box: Box) -> set[tuple[int, int]]:
        return {(r, c) for r, c in self.occupied if box.contains(r, c)}


def open_workbook(path: str | Path, *, data_only: bool = False):
    return load_workbook(path, read_only=False, data_only=data_only)


def iter_target_sheets(workbook, sheet_name: str | None = None):
    if sheet_name:
        yield workbook[sheet_name]
    else:
        for ws in workbook.worksheets:
            yield ws


def is_non_empty(value: Any) -> bool:
    return value is not None and value != ""


def is_numeric_value(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def border_side_score(side: Any, strong_styles: set[str]) -> tuple[float, float]:
    style = getattr(side, "style", None)
    if not style:
        return 0.0, 0.0
    return 1.0, 1.0 if style in strong_styles else 0.0


def cell_border_scores(cell: Any, strong_styles: set[str]) -> tuple[float, float]:
    if cell is None:
        return 0.0, 0.0
    border = getattr(cell, "border", None)
    if border is None:
        return 0.0, 0.0
    total = 0.0
    strong = 0.0
    for side_name in ("left", "right", "top", "bottom"):
        s, st = border_side_score(getattr(border, side_name), strong_styles)
        total += s
        strong += st
    return total / 4.0, strong / 4.0


def fill_key(cell: Any) -> str | None:
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type is None:
        return None
    fg = getattr(fill, "fgColor", None)
    return getattr(fg, "rgb", None) or getattr(fg, "indexed", None) or getattr(fg, "theme", None)


def alignment_key(cell: Any) -> str | None:
    alignment = getattr(cell, "alignment", None)
    if not alignment:
        return None
    parts = [alignment.horizontal, alignment.vertical]
    return "/".join([p for p in parts if p]) or None


def merged_boxes(ws: Worksheet) -> list[Box]:
    boxes: list[Box] = []
    for rng in ws.merged_cells.ranges:
        boxes.append(Box(rng.min_row, rng.min_col, rng.max_row, rng.max_col))
    return boxes


def _column_width_to_pixels_for_signal(width: float | None) -> int:
    width = 8.43 if width is None else float(width)
    return max(4, int(width * 7 + 5))


def _row_height_to_pixels_for_signal(height_pt: float | None) -> int:
    height_pt = 15.0 if height_pt is None else float(height_pt)
    return max(4, int(height_pt * 96 / 72))


def _image_extent_pixels(img: Any) -> tuple[int, int]:
    anchor = getattr(img, "anchor", None)
    ext = getattr(anchor, "ext", None)
    if ext is not None:
        width = int(getattr(ext, "cx", 0) / 9525)
        height = int(getattr(ext, "cy", 0) / 9525)
        if width > 0 and height > 0:
            return width, height
    return int(getattr(img, "width", 64) or 64), int(getattr(img, "height", 20) or 20)


def image_boxes(ws: Worksheet) -> list[Box]:
    boxes: list[Box] = []
    for img in getattr(ws, "_images", []):
        anchor = getattr(img, "anchor", None)
        if anchor is None or not hasattr(anchor, "_from"):
            continue
        start = anchor._from
        min_row = int(start.row) + 1
        min_col = int(start.col) + 1

        end_marker = None
        if hasattr(anchor, "_to") and getattr(anchor, "_to", None) is not None:
            end_marker = anchor._to
        elif hasattr(anchor, "to") and getattr(anchor, "to", None) is not None:
            end_marker = anchor.to

        if end_marker is not None:
            max_row = max(min_row, int(end_marker.row) + 1)
            max_col = max(min_col, int(end_marker.col) + 1)
            boxes.append(Box(min_row, min_col, max_row, max_col))
            continue

        width_px, height_px = _image_extent_pixels(img)
        remaining_w = width_px
        max_col = min_col
        while remaining_w > 0 and max_col <= 16384:
            letter = get_column_letter(max_col)
            width = ws.column_dimensions[letter].width if letter in ws.column_dimensions else None
            remaining_w -= _column_width_to_pixels_for_signal(width)
            if remaining_w > 0:
                max_col += 1

        remaining_h = height_px
        max_row = min_row
        while remaining_h > 0 and max_row <= 1048576:
            height = ws.row_dimensions[max_row].height if max_row in ws.row_dimensions else None
            remaining_h -= _row_height_to_pixels_for_signal(height)
            if remaining_h > 0:
                max_row += 1

        boxes.append(Box(min_row, min_col, max_row, max_col))
    return boxes


def all_image_boxes(ws: Worksheet, config: dict[str, Any]) -> list[Box]:
    workbook_path = config.get("workbook_path")
    raw_boxes: list[Box] = []
    if workbook_path and config.get("include_grouped_drawing_images", True):
        try:
            raw_boxes = drawing_image_boxes(workbook_path, ws.title, ws)
        except Exception:
            raw_boxes = []

    # Prefer raw DrawingML boxes when available. openpyxl flattens grouped drawings
    # into a single object, while the raw parser can expose each child picture.
    boxes = raw_boxes if raw_boxes else image_boxes(ws)

    # Dedupe by exact cell range while preserving order.
    seen: set[tuple[int, int, int, int]] = set()
    deduped: list[Box] = []
    for box in boxes:
        key = (box.min_row, box.min_col, box.max_row, box.max_col)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(box)
    return deduped


def effective_bounds(ws: Worksheet, config: dict[str, Any]) -> Box | None:
    rows: list[int] = []
    cols: list[int] = []

    for (row, col), cell in ws._cells.items():
        if is_non_empty(cell.value):
            rows.append(row)
            cols.append(col)

    for box in merged_boxes(ws):
        top_left = ws.cell(box.min_row, box.min_col)
        if is_non_empty(top_left.value):
            rows.extend([box.min_row, box.max_row])
            cols.extend([box.min_col, box.max_col])

    for box in all_image_boxes(ws, config):
        rows.extend([box.min_row, box.max_row])
        cols.extend([box.min_col, box.max_col])

    if not rows or not cols:
        return None

    pad_r = int(config.get("bounds_padding_rows", 0))
    pad_c = int(config.get("bounds_padding_cols", 0))
    return Box(
        max(1, min(rows) - pad_r),
        max(1, min(cols) - pad_c),
        min(1048576, max(rows) + pad_r),
        min(16384, max(cols) + pad_c),
    )


def extract_sheet_signals(ws: Worksheet, config: dict[str, Any]) -> SheetSignals:
    strong_styles = set(config.get("strong_border_styles", ["medium", "thick", "double"]))
    bounds = effective_bounds(ws, config)
    if bounds is None:
        return SheetSignals(ws.title, None, {}, set(), [], [])

    cells: dict[tuple[int, int], CellSignal] = {}
    occupied: set[tuple[int, int]] = set()

    def ensure_signal(row: int, col: int) -> CellSignal:
        key = (row, col)
        if key not in cells:
            cell = ws.cell(row, col)
            value = cell.value
            border_score, strong_border_score = cell_border_scores(cell, strong_styles)
            cells[key] = CellSignal(
                row=row,
                col=col,
                value=value,
                has_value=is_non_empty(value),
                is_numeric=is_numeric_value(value),
                is_formula=isinstance(value, str) and value.startswith("="),
                bold=bool(getattr(cell.font, "bold", False)),
                fill=str(fill_key(cell)) if fill_key(cell) is not None else None,
                alignment=alignment_key(cell),
                border_score=border_score,
                strong_border_score=strong_border_score,
            )
        return cells[key]

    if config.get("include_values", True):
        for (row, col), cell in ws._cells.items():
            if not bounds.contains(row, col):
                continue
            sig = ensure_signal(row, col)
            if sig.has_value:
                occupied.add((row, col))

    m_boxes = merged_boxes(ws) if config.get("include_merged_cells", True) else []
    for box in m_boxes:
        top_left = ws.cell(box.min_row, box.min_col)
        if not is_non_empty(top_left.value):
            continue
        for row in range(max(bounds.min_row, box.min_row), min(bounds.max_row, box.max_row) + 1):
            for col in range(max(bounds.min_col, box.min_col), min(bounds.max_col, box.max_col) + 1):
                sig = ensure_signal(row, col)
                sig.is_merged = True
                occupied.add((row, col))

    img_boxes = all_image_boxes(ws, config) if config.get("include_images", True) else []
    for box in img_boxes:
        for row in range(max(bounds.min_row, box.min_row), min(bounds.max_row, box.max_row) + 1):
            for col in range(max(bounds.min_col, box.min_col), min(bounds.max_col, box.max_col) + 1):
                sig = ensure_signal(row, col)
                sig.has_image = True
                occupied.add((row, col))

    if config.get("include_bordered_empty_cells", False):
        for row in range(bounds.min_row, bounds.max_row + 1):
            for col in range(bounds.min_col, bounds.max_col + 1):
                sig = ensure_signal(row, col)
                if sig.strong_border_score > 0:
                    occupied.add((row, col))

    # Keep style-bearing cells only if they are near effective bounds and useful for profiles.
    for (row, col), cell in ws._cells.items():
        if bounds.contains(row, col):
            ensure_signal(row, col)

    return SheetSignals(ws.title, bounds, cells, occupied, m_boxes, img_boxes)


def region_features(signals: SheetSignals, box: Box) -> dict[str, Any]:
    coords = [(r, c) for (r, c), sig in signals.cells.items() if box.contains(r, c)]
    occupied = [(r, c) for (r, c) in signals.occupied if box.contains(r, c)]
    if not coords:
        coords = [(r, c) for r in range(box.min_row, box.max_row + 1) for c in range(box.min_col, box.max_col + 1)]

    values = [signals.cells.get((r, c)) for (r, c) in coords]
    values = [v for v in values if v is not None]
    non_empty = [v for v in values if v.has_value]
    numeric = [v for v in non_empty if v.is_numeric]
    formulas = [v for v in non_empty if v.is_formula]
    text_like = [v for v in non_empty if isinstance(v.value, str) and not v.is_formula]
    bold = [v for v in values if v.bold]
    merged = [v for v in values if v.is_merged]
    images = [v for v in values if v.has_image]
    border_avg = sum(v.border_score for v in values) / max(len(values), 1)
    strong_border_avg = sum(v.strong_border_score for v in values) / max(len(values), 1)

    fill_counter = Counter(v.fill for v in values if v.fill)
    align_counter = Counter(v.alignment for v in values if v.alignment)

    return {
        "cell_area": box.area,
        "occupied_count": len(occupied),
        "occupied_density": round(len(occupied) / max(box.area, 1), 4),
        "non_empty_count": len(non_empty),
        "non_empty_density": round(len(non_empty) / max(box.area, 1), 4),
        "numeric_count": len(numeric),
        "numeric_density": round(len(numeric) / max(len(non_empty), 1), 4),
        "formula_count": len(formulas),
        "formula_density": round(len(formulas) / max(len(non_empty), 1), 4),
        "text_count": len(text_like),
        "text_density": round(len(text_like) / max(len(non_empty), 1), 4),
        "bold_count": len(bold),
        "bold_density": round(len(bold) / max(len(values), 1), 4),
        "merged_count": len(merged),
        "image_count": len(images),
        "border_coverage": round(border_avg, 4),
        "strong_border_coverage": round(strong_border_avg, 4),
        "dominant_fill": fill_counter.most_common(1)[0][0] if fill_counter else None,
        "dominant_alignment": align_counter.most_common(1)[0][0] if align_counter else None,
    }


def row_merge_signature(signals: SheetSignals, row: int, min_col: int, max_col: int) -> tuple[tuple[int, int], ...]:
    ranges: list[tuple[int, int]] = []
    for box in signals.merged_ranges:
        if box.min_row <= row <= box.max_row and not (box.max_col < min_col or box.min_col > max_col):
            ranges.append((max(min_col, box.min_col), min(max_col, box.max_col)))
    return tuple(sorted(ranges))
