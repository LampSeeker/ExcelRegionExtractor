from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from PIL import Image, ImageDraw, ImageFont
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .io import ensure_dir
from .raw_drawing import extract_drawing_images, drawing_image_pixel_box


EMU_PER_PIXEL = 9525
DEFAULT_COL_WIDTH = 8.43
DEFAULT_ROW_HEIGHT_PT = 15.0


def _font(size: int = 13, bold: bool = False, font_path: str | None = None):
    explicit = [font_path] if font_path else []
    candidates = explicit + (
        [
            "C:/Windows/Fonts/malgunbd.ttf",
            "C:/Windows/Fonts/malgun.ttf",
            "/mnt/c/Windows/Fonts/malgunbd.ttf",
            "/mnt/c/Windows/Fonts/malgun.ttf",
            "/System/Library/Fonts/AppleSDGothicNeo.ttc",
            "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
            "/usr/share/fonts/truetype/nanum/NanumBarunGothicBold.ttf",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        ]
        if bold
        else [
            "C:/Windows/Fonts/malgun.ttf",
            "C:/Windows/Fonts/malgunbd.ttf",
            "/mnt/c/Windows/Fonts/malgun.ttf",
            "/mnt/c/Windows/Fonts/malgunbd.ttf",
            "/System/Library/Fonts/AppleSDGothicNeo.ttc",
            "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
            "/usr/share/fonts/truetype/nanum/NanumBarunGothic.ttf",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
    )
    for path in candidates:
        if not path:
            continue
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            pass
    return ImageFont.load_default()


def _value_preview(value: Any, max_len: int = 42) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 1] + "…"


def _bounds_from_regions(regions: list[dict[str, Any]]) -> tuple[int, int, int, int] | None:
    if not regions:
        return None
    return (
        min(int(r["min_row"]) for r in regions),
        min(int(r["min_col"]) for r in regions),
        max(int(r["max_row"]) for r in regions),
        max(int(r["max_col"]) for r in regions),
    )


def _expand_bounds(bounds: tuple[int, int, int, int], pad_rows: int, pad_cols: int) -> tuple[int, int, int, int]:
    min_row, min_col, max_row, max_col = bounds
    return (
        max(1, min_row - pad_rows),
        max(1, min_col - pad_cols),
        min(1048576, max_row + pad_rows),
        min(16384, max_col + pad_cols),
    )


def _column_width_to_pixels(width: float | None) -> int:
    # Excel's exact conversion depends on the default font. This approximation is close enough for overlay debugging.
    width = DEFAULT_COL_WIDTH if width is None else float(width)
    if width <= 0:
        return 0
    return max(4, int(width * 7 + 5))


def _row_height_to_pixels(height_pt: float | None) -> int:
    height_pt = DEFAULT_ROW_HEIGHT_PT if height_pt is None else float(height_pt)
    if height_pt <= 0:
        return 0
    return max(4, int(height_pt * 96 / 72))


def _hex_to_rgb(value: Any, default: tuple[int, int, int] = (255, 255, 255)) -> tuple[int, int, int]:
    if value is None:
        return default
    text = str(value)
    if len(text) == 8:
        text = text[2:]
    if len(text) != 6:
        return default
    try:
        return tuple(int(text[i : i + 2], 16) for i in (0, 2, 4))  # type: ignore[return-value]
    except Exception:
        return default


def _cell_fill_rgb(cell: Any) -> tuple[int, int, int]:
    fill = getattr(cell, "fill", None)
    if fill is None or fill.fill_type is None:
        return (255, 255, 255)
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return (255, 255, 255)
    if getattr(fg, "type", None) == "rgb" and getattr(fg, "rgb", None):
        return _hex_to_rgb(fg.rgb, (255, 255, 255))
    # Theme/indexed colors are deliberately kept neutral rather than guessed incorrectly.
    return (248, 248, 238)


def _border_color(side: Any) -> tuple[int, int, int]:
    color = getattr(side, "color", None)
    if color is not None and getattr(color, "type", None) == "rgb" and getattr(color, "rgb", None):
        return _hex_to_rgb(color.rgb, (0, 0, 0))
    return (0, 0, 0)


def _border_width(style: str | None) -> int:
    if not style:
        return 0
    if style in {"medium", "mediumDashDot", "mediumDashed", "mediumDashDotDot"}:
        return 2
    if style in {"thick", "double"}:
        return 3
    return 1


def _merged_lookup(ws: Worksheet) -> tuple[dict[tuple[int, int], Any], set[tuple[int, int]]]:
    top_left_by_cell: dict[tuple[int, int], Any] = {}
    covered_non_topleft: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        top_left = (rng.min_row, rng.min_col)
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                top_left_by_cell[(row, col)] = rng
                if (row, col) != top_left:
                    covered_non_topleft.add((row, col))
    return top_left_by_cell, covered_non_topleft


def _alignment_offsets(cell: Any, rect: tuple[int, int, int, int], text_size: tuple[int, int]) -> tuple[int, int]:
    x0, y0, x1, y1 = rect
    tw, th = text_size
    align = getattr(cell, "alignment", None)
    horizontal = getattr(align, "horizontal", None) if align else None
    vertical = getattr(align, "vertical", None) if align else None
    if horizontal == "center":
        x = x0 + max(2, (x1 - x0 - tw) // 2)
    elif horizontal == "right":
        x = x1 - tw - 4
    else:
        x = x0 + 4
    if vertical == "center":
        y = y0 + max(2, (y1 - y0 - th) // 2)
    elif vertical == "bottom":
        y = y1 - th - 3
    else:
        y = y0 + 3
    return x, y


def _text_bbox(draw: ImageDraw.ImageDraw, xy: tuple[int, int], text: str, font: ImageFont.ImageFont) -> tuple[int, int, int, int]:
    try:
        return draw.textbbox(xy, text, font=font)
    except Exception:
        w, h = draw.textsize(text, font=font)
        return (xy[0], xy[1], xy[0] + w, xy[1] + h)


def _draw_wrapped_text(draw: ImageDraw.ImageDraw, rect: tuple[int, int, int, int], text: str, font: ImageFont.ImageFont, fill: tuple[int, int, int], cell: Any) -> None:
    if not text:
        return
    x0, y0, x1, y1 = rect
    max_width = max(1, x1 - x0 - 8)
    lines: list[str] = []
    for raw_line in text.split("\n"):
        line = ""
        for ch in raw_line:
            trial = line + ch
            bbox = _text_bbox(draw, (0, 0), trial, font)
            if bbox[2] - bbox[0] > max_width and line:
                lines.append(line)
                line = ch
            else:
                line = trial
        lines.append(line)
    if len(lines) > 4:
        lines = lines[:4]
        lines[-1] = lines[-1][: max(0, len(lines[-1]) - 1)] + "…"
    line_height = max(10, _text_bbox(draw, (0, 0), "가", font)[3] + 2)
    block_height = line_height * len(lines)
    first_rect = (x0, y0, x1, y1)
    tx, ty = _alignment_offsets(cell, first_rect, (min(max_width, max((_text_bbox(draw, (0, 0), l, font)[2] for l in lines), default=0)), block_height))
    for line in lines:
        draw.text((tx, ty), line, fill=fill, font=font)
        ty += line_height
        if ty > y1 - 4:
            break


def _image_bounds(img: Any, left_gutter: int, top_gutter: int, col_x: dict[int, int], row_y: dict[int, int], min_row: int, min_col: int) -> tuple[int, int, int, int] | None:
    anchor = getattr(img, "anchor", None)
    if anchor is None or not hasattr(anchor, "_from"):
        return None
    start = anchor._from
    start_row = int(start.row) + 1
    start_col = int(start.col) + 1
    if start_row not in row_y or start_col not in col_x:
        # The image may start just outside the current viewport.
        if start_row < min_row or start_col < min_col:
            return None
    x0 = left_gutter + col_x.get(start_col, col_x.get(min_col, 0)) + int(getattr(start, "colOff", 0) / EMU_PER_PIXEL)
    y0 = top_gutter + row_y.get(start_row, row_y.get(min_row, 0)) + int(getattr(start, "rowOff", 0) / EMU_PER_PIXEL)

    width = int(getattr(img, "width", 0) or 0)
    height = int(getattr(img, "height", 0) or 0)
    ext = getattr(anchor, "ext", None)
    if ext is not None:
        width = int(getattr(ext, "cx", 0) / EMU_PER_PIXEL) or width
        height = int(getattr(ext, "cy", 0) / EMU_PER_PIXEL) or height
    end_marker = None
    if hasattr(anchor, "_to") and getattr(anchor, "_to", None) is not None:
        end_marker = anchor._to
    elif hasattr(anchor, "to") and getattr(anchor, "to", None) is not None:
        end_marker = anchor.to

    if end_marker is not None:
        end_row = int(end_marker.row) + 1
        end_col = int(end_marker.col) + 1
        if end_row in row_y and end_col in col_x:
            x1 = left_gutter + col_x[end_col] + int(getattr(end_marker, "colOff", 0) / EMU_PER_PIXEL)
            y1 = top_gutter + row_y[end_row] + int(getattr(end_marker, "rowOff", 0) / EMU_PER_PIXEL)
            width = max(1, x1 - x0)
            height = max(1, y1 - y0)
    return x0, y0, x0 + max(1, width), y0 + max(1, height)


def _paste_worksheet_images(
    base: Image.Image,
    ws: Worksheet,
    left_gutter: int,
    top_gutter: int,
    col_x: dict[int, int],
    row_y: dict[int, int],
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
    workbook_path: str | Path | None = None,
) -> None:
    # Prefer raw DrawingML parsing when the xlsx path is available. openpyxl misses
    # pictures nested inside grouped shapes (`xdr:grpSp`), which is exactly how some
    # engineering spreadsheets store multiple diagram images.
    if workbook_path:
        try:
            raw_images = extract_drawing_images(workbook_path, ws.title)
        except Exception:
            raw_images = []
        if raw_images:
            for raw_img in raw_images:
                try:
                    rx0, ry0, rx1, ry1 = drawing_image_pixel_box(raw_img, ws, col_x, row_y)
                    x0 = int(left_gutter + rx0)
                    y0 = int(top_gutter + ry0)
                    x1 = int(left_gutter + rx1)
                    y1 = int(top_gutter + ry1)
                except Exception:
                    continue
                if x1 < left_gutter or y1 < top_gutter:
                    continue
                if x0 > base.width or y0 > base.height:
                    continue
                try:
                    pil_img = Image.open(BytesIO(raw_img.data)).convert("RGBA")
                except Exception:
                    continue
                target_w = max(1, x1 - x0)
                target_h = max(1, y1 - y0)
                pil_img = pil_img.resize((target_w, target_h))
                base.alpha_composite(pil_img, (x0, y0))
            return

    # Fallback: openpyxl's public image list. This works for simple pictures but not
    # for grouped drawing children.
    for img in getattr(ws, "_images", []):
        bounds = _image_bounds(img, left_gutter, top_gutter, col_x, row_y, min_row, min_col)
        if bounds is None:
            continue
        x0, y0, x1, y1 = bounds
        if x1 < left_gutter or y1 < top_gutter:
            continue
        if x0 > base.width or y0 > base.height:
            continue
        try:
            data = img._data()
            pil_img = Image.open(BytesIO(data)).convert("RGBA")
        except Exception:
            continue
        target_w = max(1, x1 - x0)
        target_h = max(1, y1 - y0)
        pil_img = pil_img.resize((target_w, target_h))
        base.alpha_composite(pil_img, (x0, y0))


def render_region_overlay(
    ws: Worksheet,
    regions: list[dict[str, Any]],
    out_path: str | Path,
    *,
    title: str | None = None,
    bounds: dict[str, Any] | None = None,
    max_rows: int = 120,
    max_cols: int = 50,
    pad_rows: int = 2,
    pad_cols: int = 2,
    preserve_dimensions: bool = True,
    include_images: bool = True,
    include_cell_text: bool = True,
    include_merged_cells: bool = True,
    scale: float = 1.0,
    font_path: str | None = None,
    workbook_path: str | Path | None = None,
) -> Path:
    """Render a worksheet-like PNG and draw region boxes on top.

    The renderer is still a debugging view, not a full Excel renderer, but it preserves the important
    signals for region review: row/column dimensions, merged cells, fills, borders, images, Korean
    fonts, and cached formula result values when the worksheet was opened with data_only=True.
    """
    out_path = Path(out_path)
    ensure_dir(out_path.parent)

    if bounds:
        raw_bounds = (
            int(bounds["min_row"]),
            int(bounds["min_col"]),
            int(bounds["max_row"]),
            int(bounds["max_col"]),
        )
    else:
        region_bounds = _bounds_from_regions(regions)
        raw_bounds = region_bounds or (1, 1, min(ws.max_row, 30), min(ws.max_column, 12))

    min_row, min_col, max_row, max_col = _expand_bounds(raw_bounds, pad_rows, pad_cols)
    if max_row - min_row + 1 > max_rows:
        max_row = min_row + max_rows - 1
    if max_col - min_col + 1 > max_cols:
        max_col = min_col + max_cols - 1

    rows = list(range(min_row, max_row + 1))
    cols = list(range(min_col, max_col + 1))
    left_gutter = 54
    top_gutter = 48

    col_widths: dict[int, int] = {}
    for col in cols:
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width if letter in ws.column_dimensions else None
        px = _column_width_to_pixels(width) if preserve_dimensions else 72
        col_widths[col] = max(4, int(px * scale))

    row_heights: dict[int, int] = {}
    for row in rows:
        height = ws.row_dimensions[row].height if row in ws.row_dimensions else None
        px = _row_height_to_pixels(height) if preserve_dimensions else 24
        row_heights[row] = max(4, int(px * scale))

    col_x: dict[int, int] = {}
    cursor = 0
    for col in cols:
        col_x[col] = cursor
        cursor += col_widths[col]
    row_y: dict[int, int] = {}
    cursor_y = 0
    for row in rows:
        row_y[row] = cursor_y
        cursor_y += row_heights[row]

    width = left_gutter + sum(col_widths.values()) + 24
    height = top_gutter + sum(row_heights.values()) + 34
    image = Image.new("RGBA", (width, height), "white")
    draw = ImageDraw.Draw(image)
    font_cache: dict[tuple[int, bool], ImageFont.ImageFont] = {}

    def get_font(size: int, bold: bool = False):
        size = max(8, int(size * scale))
        key = (size, bold)
        if key not in font_cache:
            font_cache[key] = _font(size, bold, font_path)
        return font_cache[key]

    header_font = get_font(10)
    title_font = get_font(13, True)
    draw.text((8, 8), title or ws.title, fill=(20, 20, 20), font=title_font)

    # Headers.
    for col in cols:
        x0 = left_gutter + col_x[col]
        x1 = x0 + col_widths[col]
        draw.rectangle((x0, top_gutter - 22, x1, top_gutter), fill=(245, 245, 245), outline=(210, 210, 210))
        draw.text((x0 + 4, top_gutter - 18), get_column_letter(col), fill=(80, 80, 80), font=header_font)
    for row in rows:
        y0 = top_gutter + row_y[row]
        y1 = y0 + row_heights[row]
        draw.rectangle((0, y0, left_gutter, y1), fill=(245, 245, 245), outline=(210, 210, 210))
        draw.text((8, y0 + 3), str(row), fill=(80, 80, 80), font=header_font)

    merged_by_cell, merged_non_topleft = _merged_lookup(ws)

    # Cell fills and grid. Merged ranges are drawn as one block so the visual structure matches Excel better.
    drawn_merged: set[str] = set()
    for row in rows:
        for col in cols:
            if include_merged_cells and (row, col) in merged_non_topleft:
                continue
            cell = ws.cell(row, col)
            x0 = left_gutter + col_x[col]
            y0 = top_gutter + row_y[row]
            x1 = x0 + col_widths[col]
            y1 = y0 + row_heights[row]
            draw_cell = cell
            merged_range = merged_by_cell.get((row, col)) if include_merged_cells else None
            if merged_range is not None:
                key = str(merged_range.coord)
                if key in drawn_merged:
                    continue
                drawn_merged.add(key)
                m_min_row = max(min_row, merged_range.min_row)
                m_min_col = max(min_col, merged_range.min_col)
                m_max_row = min(max_row, merged_range.max_row)
                m_max_col = min(max_col, merged_range.max_col)
                x0 = left_gutter + col_x[m_min_col]
                y0 = top_gutter + row_y[m_min_row]
                x1 = left_gutter + col_x[m_max_col] + col_widths[m_max_col]
                y1 = top_gutter + row_y[m_max_row] + row_heights[m_max_row]
                draw_cell = ws.cell(merged_range.min_row, merged_range.min_col)

            fill = _cell_fill_rgb(draw_cell)
            draw.rectangle((x0, y0, x1, y1), fill=fill, outline=(226, 226, 226))

    # Paste images before text and region overlays so boxes remain visible.
    if include_images:
        _paste_worksheet_images(image, ws, left_gutter, top_gutter, col_x, row_y, min_row, min_col, max_row, max_col, workbook_path)
        draw = ImageDraw.Draw(image)

    # Text layer.
    if include_cell_text:
        drawn_merged.clear()
        for row in rows:
            for col in cols:
                if include_merged_cells and (row, col) in merged_non_topleft:
                    continue
                cell = ws.cell(row, col)
                text = _value_preview(cell.value)
                if not text:
                    continue
                x0 = left_gutter + col_x[col]
                y0 = top_gutter + row_y[row]
                x1 = x0 + col_widths[col]
                y1 = y0 + row_heights[row]
                draw_cell = cell
                merged_range = merged_by_cell.get((row, col)) if include_merged_cells else None
                if merged_range is not None:
                    key = str(merged_range.coord)
                    if key in drawn_merged:
                        continue
                    drawn_merged.add(key)
                    m_min_row = max(min_row, merged_range.min_row)
                    m_min_col = max(min_col, merged_range.min_col)
                    m_max_row = min(max_row, merged_range.max_row)
                    m_max_col = min(max_col, merged_range.max_col)
                    x0 = left_gutter + col_x[m_min_col]
                    y0 = top_gutter + row_y[m_min_row]
                    x1 = left_gutter + col_x[m_max_col] + col_widths[m_max_col]
                    y1 = top_gutter + row_y[m_max_row] + row_heights[m_max_row]
                    draw_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                    text = _value_preview(draw_cell.value)
                font_size = int(getattr(draw_cell.font, "sz", None) or 10)
                font = get_font(font_size, bool(getattr(draw_cell.font, "bold", False)))
                color = (40, 40, 40)
                if getattr(draw_cell.font, "color", None) is not None:
                    fc = draw_cell.font.color
                    if getattr(fc, "type", None) == "rgb" and getattr(fc, "rgb", None):
                        color = _hex_to_rgb(fc.rgb, color)
                _draw_wrapped_text(draw, (x0, y0, x1, y1), text, font, color, draw_cell)

    # Border layer.
    #
    # Merged cells must be handled as a single rectangle.
    # The previous renderer drew borders by iterating every physical cell, so internal
    # grid/border lines inside merged ranges were drawn again. That made Excel merged
    # areas look like they were split into multiple cells. Here we first draw normal
    # non-merged cells, then draw only the outer perimeter of each merged range.
    def draw_border_segment(side: Any, coords: tuple[int, int, int, int]) -> None:
        width_px = _border_width(getattr(side, "style", None))
        if width_px:
            draw.line(coords, fill=_border_color(side), width=width_px)

    def cell_rect(row: int, col: int) -> tuple[int, int, int, int]:
        x0 = left_gutter + col_x[col]
        y0 = top_gutter + row_y[row]
        x1 = x0 + col_widths[col]
        y1 = y0 + row_heights[row]
        return x0, y0, x1, y1

    merged_cells_all = set(merged_by_cell.keys()) if include_merged_cells else set()

    # 1) Draw borders for normal cells only.
    for row in rows:
        for col in cols:
            if include_merged_cells and (row, col) in merged_cells_all:
                continue
            cell = ws.cell(row, col)
            x0, y0, x1, y1 = cell_rect(row, col)
            border = getattr(cell, "border", None)
            if border is None:
                continue
            draw_border_segment(border.left, (x0, y0, x0, y1))
            draw_border_segment(border.right, (x1, y0, x1, y1))
            draw_border_segment(border.top, (x0, y0, x1, y0))
            draw_border_segment(border.bottom, (x0, y1, x1, y1))

    # 2) Draw only merged range outer borders.
    if include_merged_cells:
        for rng in ws.merged_cells.ranges:
            if rng.max_row < min_row or rng.min_row > max_row or rng.max_col < min_col or rng.min_col > max_col:
                continue

            m_min_row = max(min_row, rng.min_row)
            m_min_col = max(min_col, rng.min_col)
            m_max_row = min(max_row, rng.max_row)
            m_max_col = min(max_col, rng.max_col)

            x0 = left_gutter + col_x[m_min_col]
            y0 = top_gutter + row_y[m_min_row]
            x1 = left_gutter + col_x[m_max_col] + col_widths[m_max_col]
            y1 = top_gutter + row_y[m_max_row] + row_heights[m_max_row]

            # Excel often stores merged-range borders on the edge cells, not always
            # only on the top-left cell. Use the first edge cell that has a style.
            def first_styled_side(cells: list[Any], side_name: str) -> Any:
                fallback = getattr(cells[0].border, side_name)
                for c in cells:
                    side = getattr(c.border, side_name)
                    if getattr(side, "style", None):
                        return side
                return fallback

            top_cells = [ws.cell(rng.min_row, c) for c in range(rng.min_col, rng.max_col + 1)]
            bottom_cells = [ws.cell(rng.max_row, c) for c in range(rng.min_col, rng.max_col + 1)]
            left_cells = [ws.cell(r, rng.min_col) for r in range(rng.min_row, rng.max_row + 1)]
            right_cells = [ws.cell(r, rng.max_col) for r in range(rng.min_row, rng.max_row + 1)]

            top_side = first_styled_side(top_cells, "top")
            bottom_side = first_styled_side(bottom_cells, "bottom")
            left_side = first_styled_side(left_cells, "left")
            right_side = first_styled_side(right_cells, "right")

            draw_border_segment(left_side, (x0, y0, x0, y1))
            draw_border_segment(right_side, (x1, y0, x1, y1))
            draw_border_segment(top_side, (x0, y0, x1, y0))
            draw_border_segment(bottom_side, (x0, y1, x1, y1))

    palette = [
        (220, 20, 60),
        (30, 144, 255),
        (34, 139, 34),
        (255, 140, 0),
        (148, 0, 211),
        (0, 139, 139),
        (178, 34, 34),
        (70, 130, 180),
    ]

    label_font = get_font(10, True)
    for idx, region in enumerate(regions):
        r1 = int(region["min_row"])
        c1 = int(region["min_col"])
        r2 = int(region["max_row"])
        c2 = int(region["max_col"])
        if r2 < min_row or r1 > max_row or c2 < min_col or c1 > max_col:
            continue
        rr1 = max(r1, min_row)
        cc1 = max(c1, min_col)
        rr2 = min(r2, max_row)
        cc2 = min(c2, max_col)
        color = palette[idx % len(palette)]
        x0 = left_gutter + col_x[cc1]
        y0 = top_gutter + row_y[rr1]
        x1 = left_gutter + col_x[cc2] + col_widths[cc2]
        y1 = top_gutter + row_y[rr2] + row_heights[rr2]
        for inset in range(3):
            draw.rectangle((x0 + inset, y0 + inset, x1 - inset, y1 - inset), outline=color)
        label = f"{region.get('id', idx + 1)} {region.get('range_ref', '')}"
        label_w = min(max(120, _text_bbox(draw, (0, 0), label, label_font)[2] + 10), max(120, x1 - x0))
        draw.rectangle((x0 + 2, y0 + 2, x0 + label_w, y0 + 22), fill=(255, 255, 255, 230), outline=color)
        draw.text((x0 + 5, y0 + 5), label[:36], fill=color, font=label_font)

    image.convert("RGB").save(out_path)
    return out_path


def render_algorithm_outputs(
    ws: Worksheet,
    sheet_data: dict[str, Any],
    sheet_out_dir: str | Path,
    algorithms: Iterable[str],
    config: dict[str, Any] | None = None,
) -> list[Path]:
    cfg = config or {}
    paths: list[Path] = []
    bounds = sheet_data.get("bounds")
    for algo in algorithms:
        data = sheet_data.get(algo, {})
        regions = data.get("regions") or []
        if not regions:
            continue
        path = Path(sheet_out_dir) / f"{algo}.png"
        paths.append(
            render_region_overlay(
                ws,
                regions,
                path,
                title=f"{ws.title} - {algo}",
                bounds=bounds,
                max_rows=int(cfg.get("max_rows", 120)),
                max_cols=int(cfg.get("max_cols", 50)),
                pad_rows=int(cfg.get("pad_rows", 2)),
                pad_cols=int(cfg.get("pad_cols", 2)),
                preserve_dimensions=bool(cfg.get("preserve_dimensions", True)),
                include_images=bool(cfg.get("include_images", True)),
                include_cell_text=bool(cfg.get("include_cell_text", True)),
                include_merged_cells=bool(cfg.get("include_merged_cells", True)),
                scale=float(cfg.get("scale", 1.0)),
                font_path=cfg.get("font_path"),
                workbook_path=cfg.get("workbook_path"),
            )
        )
    return paths
